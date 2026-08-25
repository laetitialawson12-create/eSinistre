import os
import io
import unicodedata
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from datetime import date, timedelta, time, datetime
from .auth_utils import get_profil_par_identifiant, reinitialiser_tentatives, profil_est_bloque, enregistrer_echec
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, update_session_auth_hash, authenticate
from django.contrib.auth.forms import SetPasswordForm, PasswordChangeForm
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.db.models import Q
from django.template.loader import render_to_string
from django.db import IntegrityError, transaction
from django.urls import reverse
from xhtml2pdf import pisa
from django.conf import settings
from decimal import Decimal, InvalidOperation
from .forms import (
    SinistreForm, ModifierProfilForm, AgentCreationForm, AssureAdminForm,
    DemanderComplementsForm, MarquerConformeForm, IndemnisationForm,
    ChefCreationForm, ModifierAgentAdminForm, ModifierChefAdminForm,
    ImportExcelForm, SansSuiteForm, ChequeForm, MotDePasseOublieForm,
    RegionForm, VilleForm, CommuneForm, RetraitChequeForm, AgenceForm,
    ModifierProfilAdminForm, StylePasswordChangeForm, AjouterNumeroSinistreForm,
    ModifierSinistreAdminForm, ImportSinistresForm, ImportLocalisationForm,
    DeclarationTiersForm,
)
from .models import (
    Sinistre, PieceJointe, Message, HistoriqueSinistre, EtapeSinistre,
    Assure, Agent, Agence, ChefDepartement, Quittance, Vehicule, Cheque,
    Paiement, Region, Commune, Ville, Agence
)


def get_profil_operateur(user):
    """Renvoie le profil (Agent/ChefDepartement/Administrateur) de l'utilisateur connecté,
    peu importe son rôle réel. Un chef peut ainsi agir sur tout ce qu'un
    agent peut faire, en plus de ses propres actions de validation. Un administrateur
    peut ainsi agir sur tout ce qu'un chef peut faire, en plus de ses propres actions."""
    
    return getattr(user, 'agent', None) or getattr(user, 'chef', None) or (user if user.is_staff else None)


def get_url_detail_dossier(user):
    """Renvoie le nom de l'URL de détail du dossier adaptée au rôle
    de l'utilisateur connecté (agent/chef/administrateur)."""
    if user.is_staff:
        return 'detail_sinistre_admin' 
    return 'detail_sinistre_chef' if hasattr(user, 'chef') else 'detail_sinistre_agent'


def get_retour_url(request, url_name_defaut):
    """Renvoie l'URL de la liste d'où l'utilisateur est arrivé (via ?next=),
    ou une URL de secours si absente/invalide (protège contre l'open redirect)."""
    next_url = request.GET.get('next') or request.POST.get('next')
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return next_url
    return reverse(url_name_defaut)


#-------------------------------------------------------------------
#                   AUTHENTIFICATION & REDIRECTION
#-------------------------------------------------------------------

# Regirige l'utilisateur connecté vers son tableau de bord
@login_required
def redirection_login(request):
    if request.user.is_staff:
        return redirect('accueil_admin')
    if hasattr(request.user, 'chef'):
        return redirect('accueil_chef')
    if hasattr(request.user, 'agent'):
        return redirect('accueil_agent')
    if hasattr(request.user, 'assure'):
        return redirect('accueil_assure')
    return redirect('login')


# Activation du compte assuré via le numéro de police
def activation_etape1(request):
    if request.method == 'POST':
        numero_police = request.POST.get('numero_police', '').strip()
        assure = Assure.objects.filter(numero_police=numero_police, compte_active=False).first()
        if assure:
            request.session['activation_assure_id'] = assure.id
            return redirect('activation_etape2')
        else:
            messages.error(request, "Numéro de police introuvable ou compte déjà activé.")
    return render(request, 'activation_etape1.html')


# Définition du mot de passe lors de la première activation du compte
def activation_etape2(request):
    assure_id = request.session.get('activation_assure_id')
    if not assure_id:
        return redirect('activation_etape1')
    assure = get_object_or_404(Assure, id=assure_id, compte_active=False)

    if request.method == 'POST':
        form = SetPasswordForm(assure.user, request.POST)
        if form.is_valid():
            form.save()
            assure.user.is_active = True
            assure.user.save()
            assure.compte_active = True
            assure.save()
            del request.session['activation_assure_id']
            login(request, assure.user, backend='django.contrib.auth.backends.ModelBackend')
            return redirect('politique_confidentialite')
    else:
        form = SetPasswordForm(assure.user)

    return render(request, 'activation_etape2.html', {'form': form, 'assure': assure})


# Gère l'acceptation obligatoire de la politique de confidentialité par l'assuré
@login_required
def politique_confidentialite(request):
    assure = getattr(request.user, 'assure', None)
    if not assure:
        return redirect('accueil_assure')

    if assure.politique_confidentialite_acceptee:
        return redirect('accueil_assure')

    if request.method == 'POST':
        assure.politique_confidentialite_acceptee = True
        assure.save()
        return redirect('accueil_assure')

    return render(request, 'politique_confidentialite.html')


# Gère la connexion personnalisée avec sécurité anti-brute-force (blocage temporaire).
def ma_vue_de_connexion(request):
    if request.method == 'POST':
        identifiant = request.POST.get('username')
        mot_de_passe = request.POST.get('password')
        
        # 1. Récupérer le profil pour vérifier s'il est bloqué
        profil = get_profil_par_identifiant(identifiant)
        
        if profil_est_bloque(profil):
            messages.error(request, "Votre compte est temporairement bloqué suite à 3 tentatives infructueuses. Veuillez patienter 5 minutes.")
            return render(request, 'login.html') # Remplacez 'login.html' par le nom exact de votre fichier HTML ci-dessus

        # 2. Tenter l'authentification Django
        user = authenticate(request, username=identifiant, password=mot_de_passe)
        
        if user is not None:
            login(request, user)
            reinitialiser_tentatives(profil) # Remet les compteurs à zéro en cas de succès
            return redirect('redirection_login') # Utilise votre vue de redirection existante pour dispatcher l'utilisateur
        else:
            # 3. Enregistrer l'échec en cas de mauvais mot de passe
            enregistrer_echec(identifiant)
            messages.error(request, "Identifiant ou mot de passe incorrect.")
            
    return render(request, 'login.html')


#-------------------------------------------------------------------
#                   ESPACE ASSURÉ
#-------------------------------------------------------------------


# Tableau de bord principal de l'assuré
@login_required
def accueil_assure(request):
    all_sinistres = Sinistre.objects.filter(assure=request.user)
    
    context = {
        'total': all_sinistres.count(),
        'soumis': all_sinistres.filter(statut='SOUMIS').count(),
        'en_cours': all_sinistres.filter(statut='EN_COURS').count(),
        'clotures': all_sinistres.filter(statut__in = ['CLOTURE', 'SANS_SUITE']).count(),
        'derniers': all_sinistres.order_by('-date_declaration'),
    }
    return render(request, 'accueil_assure.html', context)


# Déclarer un sinistre ou poursuivre sa déclaration
@login_required
def declarer_sinistre(request):
    # Si on arrive via une nouvelle demande (ex: lien avec ?new=1), on nettoie la session 
    # pour s'assurer qu'on crée un nouveau sinistre et qu'on ne modifie pas le précédent.
    if request.GET.get('new') == '1':
        request.session.pop('temp_sinistre_id', None)

    sinistre_id = request.session.get('temp_sinistre_id')
    sinistre_instance = None
    if sinistre_id:
        sinistre_instance = Sinistre.objects.filter(id=sinistre_id, assure=request.user).first()
        
    est_nouvelle_declaration = sinistre_instance is None
    
    if request.method == 'POST':
        form = SinistreForm(request.POST, request.FILES, user=request.user, instance=sinistre_instance)
        
        if form.is_valid():
            try:
                sinistre = form.save(commit=False)
                sinistre.assure = request.user
                sinistre.statut = 'SOUMIS'
                
                # Sauvegarde (déclenche full_clean() ; numero_sinistre reste vide pour l'instant)
                sinistre.save()
                
                # Attribution automatique de toutes les quittances valides (relation M2M, donc après le save())
                assure_profile = getattr(sinistre.assure, 'assure', None)
                if assure_profile and sinistre.date_survenance:
                    quittances_valides = Quittance.objects.filter(
                        contrat=assure_profile,
                        date_debut__lte=sinistre.date_survenance,
                        date_fin__gte=sinistre.date_survenance,
                    )
                    sinistre.quittances.set(quittances_valides)

                if not quittances_valides.exists():
                    sinistre.delete()
                    request.session.pop('temp_sinistre_id', None)
                    messages.error(
                        request,
                        "Aucune quittance valide ne couvre la date de survenance déclarée. "
                        "Votre déclaration ne peut pas être enregistrée. Vérifiez la date renseignée "
                        "ou contactez votre agence si vous pensez qu'il s'agit d'une erreur."
                    )
                    return render(request, 'declaration.html', {'form': form, 'title': 'Déclarer un sinistre'})
                
                fichiers = request.FILES.getlist('fichiers_justificatifs')
                for f in fichiers:
                    PieceJointe.objects.create(sinistre=sinistre, fichier=f)
                    
                # Historique : uniquement à la toute première déclaration
                if est_nouvelle_declaration:
                    HistoriqueSinistre.objects.create(
                        sinistre=sinistre,
                        statut='SOUMIS',
                        commentaires="Déclaration initiale du sinistre par l'assuré.",
                        auteur=request.user,
                    )
                    
                request.session['temp_sinistre_id'] = sinistre.id
                return redirect('confirmer_sinistre')
                
            except ValidationError as e:
                # Capture les erreurs levées par model.clean() ou save() et les renvoie au formulaire
                if hasattr(e, 'error_dict'):
                    for field, error_list in e.error_dict.items():
                        if field in form.fields:
                            for err in error_list:
                                form.add_error(field, err)
                        else:
                            for err in error_list:
                                form.add_error(None, err)
                else:
                    form.add_error(None, e)
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = SinistreForm(user=request.user, instance=sinistre_instance)
        
    return render(request, 'declaration.html', {'form': form, 'title': 'Déclarer un sinistre'})


# Vue permettant à un tiers de faire une déclaration de sinistre
def _resoudre_compte_et_vehicule(numero_police, immatriculation):
    """Retourne (user_assure, vehicule) à partir d'un n° de police et d'une immatriculation."""
    vehicule = None
    user_assure = None

    if immatriculation:
        vehicule = Vehicule.objects.filter(immatriculation__iexact=immatriculation).first()
        if vehicule:
            if vehicule.proprietaire:
                user_assure = vehicule.proprietaire
            elif vehicule.quittance:
                user_assure = vehicule.quittance.contrat.user

    if not user_assure and numero_police:
        assure_profile = Assure.objects.filter(numero_police__iexact=numero_police).first()
        if assure_profile:
            user_assure = assure_profile.user
            if not vehicule:
                vehicule = Vehicule.objects.filter(proprietaire=user_assure).first()
                if not vehicule:
                    v = Vehicule.objects.filter(quittance__contrat=assure_profile).first()
                    vehicule = v

    return user_assure, vehicule


def declarer_sinistre_tiers(request):
    if request.method == 'POST':
        form = DeclarationTiersForm(request.POST, request.FILES)
        if form.is_valid():
            numero_police = form.cleaned_data.get('numero_police')
            immatriculation = form.cleaned_data.get('immatriculation_recherche')
            user_assure, vehicule = _resoudre_compte_et_vehicule(numero_police, immatriculation)

            if not user_assure or not vehicule:
                messages.error(
                    request,
                    "Aucun contrat correspondant à ce n° de police ou cette immatriculation n'a été trouvé."
                )
                return render(request, 'declaration_tiers.html', {'form': form})

            try:
                sinistre = form.save(commit=False)
                sinistre.assure = user_assure
                sinistre.vehicule = vehicule
                sinistre.statut = 'SOUMIS'
                sinistre.est_declaration_tiers = True
                sinistre.nom_declarant = form.cleaned_data['nom_declarant']
                sinistre.save()

                assure_profile = getattr(user_assure, 'assure', None)
                if assure_profile and sinistre.date_survenance:
                    quittances_valides = Quittance.objects.filter(
                        contrat=assure_profile,
                        date_debut__lte=sinistre.date_survenance,
                        date_fin__gte=sinistre.date_survenance,
                    )
                    sinistre.quittances.set(quittances_valides)

                for f in request.FILES.getlist('fichiers_justificatifs'):
                    PieceJointe.objects.create(sinistre=sinistre, fichier=f)

                HistoriqueSinistre.objects.create(
                    sinistre=sinistre,
                    statut='SOUMIS',
                    commentaires=f"Déclaration initiale faite par un tiers ({sinistre.nom_declarant}).",
                    auteur=None,
                )

                messages.success(
                    request,
                    "Votre déclaration a été enregistrée. L'attestation vous sera envoyée au numéro fourni."
                )
                return redirect('declarer_sinistre_tiers')

            except ValidationError as e:
                if hasattr(e, 'error_dict'):
                    for field, error_list in e.error_dict.items():
                        for err in error_list:
                            form.add_error(field if field in form.fields else None, err)
                else:
                    form.add_error(None, e)
    else:
        form = DeclarationTiersForm()

    return render(request, 'declaration_tiers.html', {'form': form})


# Vue permettant à l'assuré de consulter son dossier.
@login_required
def detail_sinistre(request, sinistre_id):
    sinistre = get_object_or_404(Sinistre, id=sinistre_id, assure=request.user)
    nb_non_lus = sinistre.messages.exclude(auteur=request.user).filter(lu=False).count()
    
    if request.method == 'POST' and 'contenu' in request.POST:
        Message.objects.create(
            sinistre=sinistre,
            auteur=request.user,
            contenu=request.POST.get('contenu')
        )
        messages.success(request, "Message envoyé avec succès.")
        return redirect('detail_sinistre', sinistre_id=sinistre.id)

    historique_brut = list(
        sinistre.historique.exclude(commentaires__icontains="Prix").order_by('date_changement')
    )

    historique_filetre = []
    for h in historique_brut:
        if 'indemnisation' in h.commentaires.lower():
            if sinistre.indemnisation_validee:
                if sinistre.prix_retenu == 0:
                    h.commentaires = "Votre sinistre a été jugé non indemnisable par le département sinistres."
                else:
                    h.commentaires = "Indemnisation saisie et validée par le chef de département."
                historique_filetre.append(h)
        elif 'verifié par l\'agent' in h.commentaires.lower():
            h.commentaires = "Votre dossier a été vérifié et transmis au Chef de département pour validation."
            historique_filetre.append(h)
        else:
            historique_filetre.append(h)

    sinistre.messages.exclude(auteur=request.user).filter(lu=False).update(lu=True)
    
    context = {
        'sinistre': sinistre,
        'documents': sinistre.pieces.all(),
        'historique': historique_filetre,
        'discussion': sinistre.messages.all().order_by('date_envoi'),
        'nb_non_lus': nb_non_lus,
    }
    return render(request, 'detail_sinistre.html', context)


# Consulter les détails d'un dossier sinistre côté assuré (suivi, messages, historique).
@login_required
def suivi_sinistres(request):
    sinistres = Sinistre.objects.filter(assure=request.user).order_by('-date_declaration')
    return render(request, 'suivi_sinistres.html', {'sinistres': sinistres})


# Permet à l'assuré de répondre à une demande de pièces complémentaires.
@login_required
def fournir_complements(request, sinistre_id):
    sinistre = get_object_or_404(Sinistre, id=sinistre_id, assure=request.user)

    if sinistre.statut != 'ATTENTE_COMPLEMENTS':
        messages.error(request, "Ce dossier ne requiert pas de compléments actuellement.")
        return redirect('suivi_sinistres')

    if request.method == 'POST':
        sinistre.precision = request.POST.get('precisions')
        sinistre.statut = 'SOUMIS'
        sinistre.save()

        fichiers = request.FILES.getlist('documents')
        for f in fichiers:
            PieceJointe.objects.create(sinistre=sinistre, fichier=f)

        messages.success(request, f"Vos compléments pour le dossier {sinistre.numero_sinistre} ont été transmis.")
        return redirect('suivi_sinistres')

    return render(request, 'fournir_complements.html', {'sinistre': sinistre})


# Permet à l'assuré d'annuler et supprimer la déclaration de sinistre en cours
@login_required
def annuler_declaration(request):
    sinistre_id = request.session.pop('temp_sinistre_id', None)
    if sinistre_id:
        Sinistre.objects.filter(id=sinistre_id, assure=request.user).delete()
    return redirect('accueil_assure')


# Page de récapitulatif avant la validation définitive de la déclaration.
@login_required
def confirmer_sinistre(request):
    sinistre_id = request.session.get('temp_sinistre_id')
    sinistre = Sinistre.objects.filter(id=sinistre_id, assure=request.user).first() if sinistre_id else None

    if not sinistre:
        request.session.pop('temp_sinistre_id', None)
        messages.error(request, "Aucune déclaration en cours. Merci de recommencer.")
        return redirect('declarer_sinistre')

    return render(request, 'confirmation.html', {'sinistre': sinistre, 'pieces': sinistre.pieces.all()})


# Finalise la transmission du dossier de sinistre à l'agent
@login_required
def finaliser_envoi(request):
    if request.method == 'POST':
        sinistre_id = request.session.get('temp_sinistre_id')
        sinistre = get_object_or_404(Sinistre, id=sinistre_id, assure=request.user) if sinistre_id else None

        if 'temp_sinistre_id' in request.session:
            del request.session['temp_sinistre_id']

        if sinistre:
            if sinistre.quittances.exists():
                messages.success(request, f"Votre sinistre a été enregistré avec succès.")
            else:
                messages.info(
                    request,
                    f"Votre sinistre a été enregistré avec succès. "
                )
        return redirect('accueil_assure')
    return redirect('confirmer_sinistre')


# Affiche la liste de tous les documents rattachés aux sinistres de l'assuré.
@login_required
def documents_assure(request):
    pieces = PieceJointe.objects.filter(sinistre__assure=request.user)
    return render(request, 'documents_assure.html', {'pieces': pieces})


# Affiche le profil de l'assuré
@login_required
def profil_assure(request):
    return render(request, 'profil_assure.html', {'user': request.user})


# Permet à l'assuré de modifier ses informations personnelles
@login_required
def modifier_profil(request):
    assure = getattr(request.user, 'assure', None)

    if request.method == 'POST':
        form = ModifierProfilForm(request.POST)
        if form.is_valid():
            request.user.email = form.cleaned_data['email']
            request.user.save(update_fields=['email'])
            assure.telephone = form.cleaned_data['telephone']
            assure.save(update_fields=['telephone'])
            messages.success(request, "Vos informations on été mis à jour avec succès.")
            return redirect('profil_assure')
    else:
        form = ModifierProfilForm(initial={'email':request.user.email, 'telephone':assure.telephone})
    return render(request, 'modifier_profil.html', {'form':form})


# Permet à l'assuré de voir ses différents contrats
@login_required
def mes_contrats(request):
    quittances = Quittance.objects.filter(contrat__user=request.user).prefetch_related('vehicules')
    return render(request, 'mes_contrats.html', {'quittances': quittances})


# Permet à l'assuré de consulter le détail d'une quittance et ses véhicules rattachés à son contrat.
@login_required
def detail_contrat_assure(request, quittance_id):
    quittance = get_object_or_404(Quittance, id=quittance_id, contrat__user=request.user)
    vehicules = quittance.vehicules.all() if hasattr(quittance, 'vehicules') else []
    
    return render(request, 'detail_contrat_assure.html', {
        'quittance': quittance,
        'vehicules': vehicules,
    })


#-------------------------------------------------------------------
#                   ESPACE AGENT
#-------------------------------------------------------------------


# Redirige l'utilisateur vers sa page d'acceuil
@login_required
def accueil_agent(request):
    agent = getattr(request.user, 'agent', None)
    if not agent:
        return redirect('accueil_assure')

    sinistres = Sinistre.objects.all().order_by('-date_survenance')

    today = timezone.localdate()
    debut_jour = timezone.make_aware(datetime.combine(today, time.min))
    fin_jour = timezone.make_aware(datetime.combine(today, time.max))
    
    sinistres_du_jour = sinistres.filter(
        date_declaration__gte=debut_jour,
        date_declaration__lte=fin_jour,
    ).order_by('-date_declaration')
    
    
    context = {
        'agent': agent,
        'soumis': sinistres.count(),
        'a_instruire': sinistres.filter(statut='SOUMIS').count(),
        'attente_complements': sinistres.filter(statut='ATTENTE_COMPLEMENTS').count(),
        'en_cours': sinistres.filter(statut="EN_COURS").count(),
        # 'valides_ce_mois': sinistres.filter(
        #     statut='CLOTURE',
        #     date_cloture__month=timezone.now().month,
        #     date_cloture__year=timezone.now().year,
        # ).count(),
        'derniers_sinistres': sinistres_du_jour,
    }
    return render(request, 'accueil_agent.html', context)


# Tableau de bord alternatif affichant les sinistres à instruire
@login_required
def dossiers_a_valider_agent(request):
    agent = getattr(request.user, 'agent', None)
    if not agent:
        return redirect('accueil_assure')

    sinistres = Sinistre.objects.filter(
        statut__in=['SOUMIS', 'ATTENTE_COMPLEMENTS', 'A_CORRIGER'],
    ).order_by('date_declaration')

    nature = request.GET.get('nature', '')
    recherche = request.GET.get('q', '').strip()
    periode_option = request.GET.get('periode_option', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    today = timezone.localdate()
    if periode_option == 'auj':
        date_debut = date_fin = today.isoformat()
    elif periode_option == '7j':
        date_debut = (today - timedelta(days=6)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'mois':
        date_debut = today.replace(day=1).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'trimestre':
        date_debut = (today - timedelta(days=89)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'annee':
        date_debut = today.replace(month=1, day=1).isoformat()
        date_fin = today.isoformat()
        
    if nature:
        sinistres = sinistres.filter(nature=nature)
    if recherche:
        sinistres = sinistres.filter(
            Q(assure__assure__numero_police__icontains=recherche)
        )
    if date_debut:
        debut_dt = datetime.combine(date.fromisoformat(date_debut), time.min)
        sinistres = sinistres.filter(date_survenance__gte=timezone.make_aware(debut_dt))
    if date_fin:
        fin_dt = datetime.combine(date.fromisoformat(date_fin), time.max)
        sinistres = sinistres.filter(date_survenance__lte=timezone.make_aware(fin_dt))
        
    context = {
        'agent': agent,
        'sinistres': sinistres,
        'nature_choices': Sinistre.NATURE_CHOICES,
        'nature_selectionnee': nature,
        'recherche': recherche,
        'periode_option': periode_option,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    return render(request, 'dossiers_a_valider_agent.html', context)


# Permet à un agent de s'assigner la prise en charge d'un dossier soumis.
@login_required
def prendre_en_charge(request, sinistre_id):
    agent = get_profil_operateur(request.user)
    if not agent:
        return redirect('accueil_assure')

    sinistre = get_object_or_404(Sinistre, id=sinistre_id)

    if sinistre.statut == 'SOUMIS' and not sinistre.agent_traitant:
        sinistre.agent_traitant = request.user.get_full_name() or request.user.username
        sinistre.save()
        HistoriqueSinistre.objects.create(
            sinistre=sinistre,
            statut=sinistre.statut,
            commentaires="Dossier pris en charge par l'agent pour vérification.",
            auteur=request.user,
        )
        messages.success(request, f"Dossier {sinistre.numero_sinistre} pris en charge.")

    return redirect(get_url_detail_dossier(request.user), sinistre_id=sinistre.id)


# Permet à l'agent de voir les informations ou détails d'un sinistre
@login_required
def detail_sinistre_agent(request, sinistre_id):
    agent = getattr(request.user, 'agent', None)
    if not agent:
        return redirect('accueil_assure')

    sinistre = get_object_or_404(Sinistre, id=sinistre_id)
    assure_profile = getattr(sinistre.assure, 'assure', None)
    retour_url = get_retour_url(request, 'tous_sinistres_agent')

    nb_non_lus = sinistre.messages.exclude(auteur=request.user).filter(lu=False).count()
    
    # Récupération des quittances éligibles
    quittances_disponibles = Quittance.objects.none()
    if assure_profile and sinistre.date_survenance:
        date_ref = sinistre.date_survenance.date() if hasattr(sinistre.date_survenance, 'date') else sinistre.date_survenance
        quittances_disponibles = Quittance.objects.filter(
            contrat=assure_profile,
            date_debut__lte=date_ref,
            date_fin__gte=date_ref,
        ).order_by('-date_debut')

    detail_url = reverse('detail_sinistre_agent', args=[sinistre.id])

    # Traitement des soumissions POST
    if request.method == 'POST':
        # Cas 1 : Modification ou saisie du prix retenu
        if 'prix_retenu' in request.POST:
            if sinistre.statut != 'EN_COURS':
                messages.error(
                    request,
                    "Le prix retenu ne peut être saisi qu'après validation du dossier par le Chef de département."
                )
                return redirect(f"{detail_url}?next={retour_url}")

            nouveau_prix = request.POST.get('prix_retenu')
            if nouveau_prix:
                prix_existant = sinistre.prix_retenu

                sinistre.prix_retenu = nouveau_prix
                sinistre.save()

                if prix_existant is None:
                    commentaire = f"Prix retenu : {nouveau_prix} FCFA."
                else:
                    commentaire = f"Prix modifié : {nouveau_prix} FCFA (ancien prix : {prix_existant} FCFA)."

                HistoriqueSinistre.objects.create(
                    sinistre=sinistre,
                    statut=sinistre.statut,
                    commentaires=commentaire,
                    auteur=request.user,
                )
                messages.success(request, "Le prix retenu a été mis à jour avec succès.")
                return redirect(f"{detail_url}?next={retour_url}")

        # Cas 2 : Envoi d'un message direct à l'assuré
        elif 'contenu' in request.POST:
            Message.objects.create(sinistre=sinistre, auteur=request.user, contenu=request.POST.get('contenu'))
            messages.success(request, "Message transmis.")
            return redirect(f"{detail_url}?next={retour_url}")

    sinistre.messages.exclude(auteur=request.user).filter(lu=False).update(lu=True)
    context = {
        'agent': agent,
        'sinistre': sinistre,
        'retour_url': retour_url,
        'historique': sinistre.historique.all().order_by('date_changement'),
        'discussion': sinistre.messages.all().order_by('date_envoi'),
        'documents': sinistre.pieces.all(),
        'quittances_disponibles': quittances_disponibles,
        'nb_non_lus': nb_non_lus,
    }
    return render(request, 'detail_sinistre_agent.html', context)


# Permet à l'agent de marquer un dossier comme conforme et l'envoyer au chef
@login_required
def marquer_conforme(request, sinistre_id):
    operateur = get_profil_operateur(request.user)
    if not operateur:
        return redirect('accueil_assure')
    
    sinistre = get_object_or_404(Sinistre, id=sinistre_id)
    retour_url = get_retour_url(request, 'tous_sinistres_agent')
    detail_url = reverse(get_url_detail_dossier(request.user), args=[sinistre_id])

    if not sinistre.numero_sinistre:
        messages.error(request, "Le numéro de sinistre doit être ajouté avant de marquer le dossier comme conforme.")
        return redirect(f"{detail_url}?next={retour_url}")

    if request.method != 'POST':
        return redirect(f"{detail_url}?next={retour_url}")

    nature = request.POST.get('nature') or sinistre.nature
    pv_verifie = request.POST.get('pv_verifie') == 'on' if 'pv_verifie' in request.POST else sinistre.pv_verifie
    taux_responsabilite = request.POST.get('taux_responsabilite')

    if not nature:
        messages.error(request, "Veuillez renseigner la nature du sinistre.")
        return redirect(f"{detail_url}?next={retour_url}")
    if not pv_verifie:
        messages.error(request, "Le PV doit être vérifié avant l'envoi au Chef.")
        return redirect(f"{detail_url}?next={retour_url}")

    sinistre.nature = nature
    sinistre.pv_verifie = pv_verifie
    if taux_responsabilite:
        try:
            sinistre.taux_responsabilite = Decimal(taux_responsabilite)
        except InvalidOperation:
            messages.error(request, "Taux de responsabilité invalide.")
            return redirect(f"{detail_url}?next={retour_url}")

    if sinistre.statut == 'A_CORRIGER':
        sinistre.statut = 'ATTENTE_VALIDATION'
        sinistre.save()
        HistoriqueSinistre.objects.create(
            sinistre=sinistre, statut=sinistre.statut,
            commentaires=f"Prix révisé à {sinistre.prix_retenu} FCFA - Dossier transmis au chef pour validation de l'indemnisation.",
            auteur=request.user
        )
    else:
        sinistre.statut = 'ATTENTE_VALIDATION'
        sinistre.save()
        HistoriqueSinistre.objects.create(
            sinistre=sinistre, statut=sinistre.statut,
            commentaires=f"Dossier jugé conforme, numéro officiel {sinistre.numero_sinistre} attribué, transmis au Chef de département",
            auteur=request.user
        )

    messages.success(request, f"Dossier {sinistre.numero_sinistre} transmis au Chef.")
    return redirect(f"{detail_url}?next={retour_url}")


# Permet à l'agent de demander des pièces ou informations manquantes ou complémentaires
@login_required
def demander_complements(request, sinistre_id):
    agent = get_profil_operateur(request.user)
    if not agent:
        return redirect('accueil_assure')

    sinistre = get_object_or_404(Sinistre, id=sinistre_id)

    if request.method == 'POST':
        form = DemanderComplementsForm(request.POST)
        if form.is_valid():
            sinistre.statut = 'ATTENTE_COMPLEMENTS'
            sinistre.save()
            motif = form.cleaned_data['motif']
            Message.objects.create(sinistre=sinistre, auteur=request.user, contenu=motif)
            HistoriqueSinistre.objects.create(
                sinistre=sinistre,
                statut='ATTENTE_COMPLEMENTS',
                commentaires=motif,
                auteur=request.user,
            )
            messages.success(request, f"Demande de compléments envoyée pour le dossier {sinistre.numero_sinistre}.")
            return redirect(get_url_detail_dossier(request.user), sinistre_id=sinistre.id)
    else:
        form = DemanderComplementsForm()

    return render(request, 'demander_complements.html', {'sinistre': sinistre, 'form': form})


# Permet à l'agent de saisir le prix retenu par le département indemnisation après que le dossier soit passé à en cours
@login_required
def saisir_prix_retenu(request, sinistre_id):
    agent = get_profil_operateur(request.user)
    if not agent:
        return redirect('accueil_assure')

    sinistre = get_object_or_404(Sinistre, id=sinistre_id)
    retour_url = get_retour_url(request, 'tous_sinistres_agent')
    detail_url = reverse(get_url_detail_dossier(request.user), args=[sinistre_id])

    autorise = sinistre.statut == 'EN_COURS' or (
        sinistre.statut == 'A_CORRIGER' and sinistre.attestation_generee
    )

    if not autorise:
        messages.error(
            request,
            "Le prix retenu ne peut être saisi qu'après validation du dossier par le Chef de département."
        )
        return redirect(f"{detail_url}?next={retour_url}")

    if getattr(sinistre, 'indemnisation_validee', False):
        messages.error(request, "Le prix retenu a été validé par le Chef. Il ne peut plus être modifié.")
        return redirect(f"{detail_url}?next={retour_url}")

    if request.method == 'POST':
        nouveau_prix = request.POST.get('prix_retenu')
        if nouveau_prix:
            prix_existant = sinistre.prix_retenu

            sinistre.prix_retenu = nouveau_prix
            sinistre.save()

            if prix_existant is None:
                if Decimal(nouveau_prix) == 0:
                    commentaire = "Prix retenu saisi à 0 FCFA - sinistre jugé non indemnisable."
                else:
                    commentaire = f"Prix retenu à la saisie : {nouveau_prix} FCFA."
            else:
                if Decimal(nouveau_prix) == 0:
                    commentaire = f"Prix modifié : sinistre jugé non indemnisable (ancien prix : {prix_existant} FCFA)."
                else:
                    commentaire = f"Prix modifié : {nouveau_prix} FCFA (ancien prix : {prix_existant} FCFA)."

            HistoriqueSinistre.objects.create(
                sinistre=sinistre,
                statut=sinistre.statut,
                commentaires=commentaire,
                auteur=request.user,
            )
            messages.success(request, "Le prix retenu a été enrégistré avec succès.")
        return redirect(f"{detail_url}?next={retour_url}")
    
    
# Permet à l'agent d'enregistrer les détails du paiement pour un sinistre clôturé
@login_required
def saisir_indemnisation(request, sinistre_id):
    agent = get_profil_operateur(request.user)
    if not agent:
        return redirect('accueil_assure')

    sinistre = get_object_or_404(Sinistre, id=sinistre_id, statut='CLOTURE')

    if request.method == 'POST':
        form = IndemnisationForm(request.POST)
        if form.is_valid():
            for champ, valeur in form.cleaned_data.items():
                setattr(sinistre, champ, valeur)
            sinistre.save()
            HistoriqueSinistre.objects.create(
                sinistre=sinistre,
                statut='CLOTURE',
                commentaires=(
                    f"Chèque n°{sinistre.numero_cheque} ({sinistre.banque_cheque}) remis à "
                    f"{sinistre.beneficiaire_nom} {sinistre.beneficiaire_prenoms} "
                    f"({sinistre.beneficiaire_telephone})."
                ),
                auteur=request.user,
            )
            messages.success(request, "Informations d'indemnisation enregistrées.")
            return redirect(get_url_detail_dossier(request.user), sinistre_id=sinistre.id)
    else:
        form = IndemnisationForm()

    return render(request, 'saisir_indemnisation.html', {'sinistre': sinistre, 'form': form})


# Fonction affichant à l'agent les dossiers de sinistres avec le statut en cours
@login_required
def dossiers_en_cours(request):
    agent = getattr(request.user, 'agent', None)
    if not agent:
        return redirect('accueil_assure')
    sinistres = Sinistre.objects.filter(
        statut__in=['ATTENTE_VALIDATION', 'EN_COURS', 'A_CORRIGER', 'REOUVERT'],
    ).order_by('-date_declaration')
    
    nature = request.GET.get('nature', '')
    recherche = request.GET.get('q', '').strip()
    periode_option = request.GET.get('periode_option', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')

    today = timezone.localdate()
    if periode_option == 'auj':
        date_debut = date_fin = today.isoformat()
    elif periode_option == '7j':
        date_debut = (today - timedelta(days=6)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'mois':
        date_debut = today.replace(day=1).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'trimestre':
        date_debut = (today - timedelta(days=89)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'annee':
        date_debut = today.replace(month=1, day=1).isoformat()
        date_fin = today.isoformat()
    
    if nature:
        sinistres = sinistres.filter(nature=nature)
    if recherche:
        sinistres = sinistres.filter(
            Q(numero_sinistre__icontains=recherche) | Q(assure__assure__numero_police__icontains=recherche)
        )
    if date_debut:
        debut_dt = datetime.combine(date.fromisoformat(date_debut), time.min)
        sinistres = sinistres.filter(date_survenance__gte=timezone.make_aware(debut_dt))
    if date_fin:
        fin_dt = datetime.combine(date.fromisoformat(date_fin), time.max)
        sinistres = sinistres.filter(date_survenance__lte=timezone.make_aware(fin_dt))
        
    context = {
        'agent': agent,
        'sinistres': sinistres,
        'nature_choices': Sinistre.NATURE_CHOICES,
        'nature_selectionnee': nature,
        'recherche': recherche,
        'periode_option': periode_option,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    
    return render(request, 'dossiers_en_cours.html', context)


# Fonction affichant à l'agent les dossiers de sinistres avec le statut clôturé
@login_required
def dossiers_clotures(request):
    agent = getattr(request.user, 'agent', None)
    if not agent:
        return redirect('accueil_assure')
    sinistres = Sinistre.objects.filter(
        statut__in=['CLOTURE', 'SANS_SUITE'],
    ).order_by('-date_declaration')

    nature = request.GET.get('nature', '')
    recherche = request.GET.get('q', '').strip()
    periode_option = request.GET.get('periode_option', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    today = timezone.localdate()
    if periode_option == 'auj':
        date_debut = date_fin = today.isoformat()
    elif periode_option == '7j':
        date_debut = (today - timedelta(days=6)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'mois':
        date_debut = today.replace(day=1).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'trimestre':
        date_debut = (today - timedelta(days=89)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'annee':
        date_debut = today.replace(month=1, day=1).isoformat()
        date_fin = today.isoformat()
        
    if nature:
        sinistres = sinistres.filter(nature=nature)
    if recherche:
        sinistres = sinistres.filter(
            Q(numero_sinistre__icontains=recherche) | Q(assure__assure__numero_police__icontains=recherche)
        )
    if date_debut:
        debut_dt = datetime.combine(date.fromisoformat(date_debut), time.min)
        sinistres = sinistres.filter(date_survenance__gte=timezone.make_aware(debut_dt))
    if date_fin:
        fin_dt = datetime.combine(date.fromisoformat(date_fin), time.max)
        sinistres = sinistres.filter(date_survenance__lte=timezone.make_aware(fin_dt))

    context = {
        'agent': agent,
        'sinistres': sinistres,
        'nature_choices': Sinistre.NATURE_CHOICES,
        'nature_selectionnee': nature,
        'recherche': recherche,
        'periode_option': periode_option,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    return render(request, 'dossiers_clotures_agent.html', context)


# Fonction affichant à l'agent les dossiers de sinistres avec le statut à corriger
@login_required
def dossiers_a_corriger_agent(request):
    agent = getattr(request.user, 'agent', None)
    if not agent:
        return redirect('accueil_assure')
    sinistres_a_corriger = Sinistre.objects.filter(
        statut="A_CORRIGER",
    ).order_by('date_declaration')

    nature = request.GET.get('nature', '')
    recherche = request.GET.get('q', '').strip()
    periode_option = request.GET.get('periode_option', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    today = timezone.localdate()
    if periode_option == 'auj':
        date_debut = date_fin = today.isoformat()
    elif periode_option == '7j':
        date_debut = (today - timedelta(days=6)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'mois':
        date_debut = today.replace(day=1).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'trimestre':
        date_debut = (today - timedelta(days=89)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'annee':
        date_debut = today.replace(month=1, day=1).isoformat()
        date_fin = today.isoformat()
        
    if nature:
        sinistres_a_corriger = sinistres_a_corriger.filter(nature=nature)
    if recherche:
        sinistres_a_corriger = sinistres_a_corriger.filter(
            Q(numero_sinistre__icontains=recherche) | Q(assure__assure__numero_police__icontains=recherche)
        )
    if date_debut:
        debut_dt = datetime.combine(date.fromisoformat(date_debut), time.min)
        sinistres_a_corriger = sinistres_a_corriger.filter(date_survenance__gte=timezone.make_aware(debut_dt))
    if date_fin:
        fin_dt = datetime.combine(date.fromisoformat(date_fin), time.max)
        sinistres_a_corriger = sinistres_a_corriger.filter(date_survenance__lte=timezone.make_aware(fin_dt))

    context = {
        'agent': agent,
        'sinistres_a_corriger': sinistres_a_corriger,
        'nature_choices': Sinistre.NATURE_CHOICES,
        'nature_selectionnee': nature,
        'recherche': recherche,
        'periode_option': periode_option,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    return render(request, 'dossiers_a_corriger_agent.html', context)


# Fonction affichant à l'agent tous les dossiers de sinistres
@login_required
def tous_sinistres_agent(request):
    agent = getattr(request.user, 'agent', None)
    if not agent:
        return redirect('accueil_assure')

    sinistres = Sinistre.objects.select_related('assure', 'vehicule', 'region').filter(
    ).order_by('-date_declaration')

    statut = request.GET.get('statut', '')
    nature = request.GET.get('nature', '')
    recherche = request.GET.get('q', '').strip()
    periode_option = request.GET.get('periode_option', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    # Options rapides : elles écrasent la saisie manuelle si cliquées
    today = timezone.localdate()
    if periode_option == 'auj':
        date_debut = date_fin = today.isoformat()
    elif periode_option == '7j':
        date_debut = (today - timedelta(days=6)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'mois':
        date_debut = today.replace(day=1).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'trimestre':
        date_debut = (today - timedelta(days=89)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'annee':
        date_debut = today.replace(month=1, day=1).isoformat()
        date_fin = today.isoformat()
    
    if statut:
        sinistres = sinistres.filter(statut=statut)
    if nature:
        sinistres = sinistres.filter(nature=nature)
    if recherche:
        sinistres = sinistres.filter(
            Q(numero_sinistre__icontains=recherche) | Q(assure__assure__numero_police__icontains=recherche)
        )
    if date_debut:
        debut_dt = datetime.combine(date.fromisoformat(date_debut), time.min)
        sinistres = sinistres.filter(date_survenance__gte=timezone.make_aware(debut_dt))
    if date_fin:
        fin_dt = datetime.combine(date.fromisoformat(date_fin), time.max)
        sinistres = sinistres.filter(date_survenance__lte=timezone.make_aware(fin_dt))

    context = {
        'agent' : agent,
        'sinistres' : sinistres,
        'statut_choices' : Sinistre.STATUS_CHOICES,
        'nature_choices' : Sinistre.NATURE_CHOICES,
        'statut_selectionne' : statut,
        'nature_selectionnee' : nature,
        'recherche' : recherche,
        'periode_option' : periode_option,
        'date_debut' : date_debut,
        'date_fin' : date_fin,
    }

    return render(request, 'tous_sinistres_agent.html', context)


# Permet à l'agent d'ajouter le numéro du sinistre
@login_required
def ajouter_numero_sinistre(request, sinistre_id):
    sinistre = get_object_or_404(Sinistre, pk=sinistre_id)
    
    if request.method == 'POST':    
        form = AjouterNumeroSinistreForm(request.POST, instance=sinistre)
        if form.is_valid():
            form.save()
            messages.success(request, "Le numéro sinsitre a été bien modifié.")
            return redirect(get_url_detail_dossier(request.user), sinistre_id=sinistre.pk)
    else:
        form = AjouterNumeroSinistreForm(instance=sinistre)
        
    context={
        'form': form,
        'sinistre': sinistre,
        'base_layout': 'base_chef.html' if hasattr(request.user, 'chef') else 'base_agent.html',
        'url_detail_dossier': get_url_detail_dossier(request.user),
    }
    return render(request, 'ajouter_numero_sinistre.html', context)


# Permet à l'agent de modifier son mot de passe
@login_required
def changer_mot_de_passe_agent(request):
    agent = getattr(request.user, 'agent', None)
    if not agent:
        return redirect('accueil_assure')

    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, form.user)
            agent.doit_changer_mot_de_passe = False
            agent.compte_active = True
            agent.save()
            messages.success(request, "Mot de passe modifié avec succès.")
            return redirect('accueil_agent')
    else:
        form = PasswordChangeForm(request.user)

    return render(request, 'changer_mot_de_passe_agent.html', {'form': form})


# Affiche le profil de l'agent
@login_required
def profil_agent(request):
    agent = getattr(request.user, 'agent', None)
    if not agent:
        return redirect('accueil_assure')
    return render(request, 'profil_agent.html', {'agent': agent})


# Permet à l'agent de modifier certaines de ses informations personnelles
@login_required
def modifier_profil_agent(request):
    agent = getattr(request.user, 'agent', None)
    if not agent:
        return redirect('accueil_assure')
    
    if request.method == 'POST':
        form = ModifierProfilForm(request.POST)
        if form.is_valid():
            request.user.email = form.cleaned_data['email']
            request.user.save(update_fields=['email'])
            agent.telephone = form.cleaned_data['telephone']
            agent.save(update_fields=['telephone'])
            messages.success(request, "Vos informations on été mis à jour avec succès.")
            return redirect('profil_agent')
    else:
        form = ModifierProfilForm(initial={'email':request.user.email, 'telephone':agent.telephone})
    return render(request, 'modifier_profil_agent.html', {'form':form})


#-------------------------------------------------------------------
#                   ESPACE CHEF DE DÉPARTEMENT
#-------------------------------------------------------------------


# Page d'acceuil du chef
@login_required
def accueil_chef(request):
    chef = getattr(request.user, 'chef', None)
    if not chef:
        return redirect('accueil_assure')

    sinistres = Sinistre.objects.all().order_by('-date_survenance')
    
    today = timezone.localdate()
    debut_jour = timezone.make_aware(datetime.combine(today, time.min))
    fin_jour = timezone.make_aware(datetime.combine(today, time.max))
        
    sinistres_du_jour = sinistres.filter(
        date_declaration__gte=debut_jour,
        date_declaration__lte=fin_jour,
    ).order_by('-date_declaration')
        
    context = {
        'chef': chef,
        'declarer': sinistres.count(),
        'soumis': sinistres.filter(statut='SOUMIS').count(),
        'a_valider': sinistres.filter(statut='ATTENTE_VALIDATION').count(),
        'en_cours': sinistres.filter(statut='EN_COURS').count(),
        'a_corriger': sinistres.filter(statut='A_CORRIGER').count(),
        'en_attente_de_complements': sinistres.filter(statut='EN_ATTENTE_DE_COMPLEMENTS').count(),
        'clotures_ce_mois': sinistres.filter(
            statut='CLOTURE',
            date_declaration__month=timezone.now().month,
            date_declaration__year=timezone.now().year,
        ).count(),
        'derniers_sinistres': sinistres_du_jour,
    }
    return render(request, 'accueil_chef.html', context)


# Fonction affichant les dossiers au statut à valider au chef
@login_required
def dossiers_a_valider(request):
    chef = getattr(request.user, 'chef', None)
    if not chef:
        return redirect('accueil_assure')
    sinistres = Sinistre.objects.filter(
        statut='ATTENTE_VALIDATION',
    ).order_by('-date_declaration')
    
    nature = request.GET.get('nature', '')
    recherche = request.GET.get('q', '').strip()
    periode_option = request.GET.get('periode_option', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')

    today = timezone.localdate()
    if periode_option == 'auj':
        date_debut = date_fin = today.isoformat()
    elif periode_option == '7j':
        date_debut = (today - timedelta(days=6)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'mois':
        date_debut = today.replace(day=1).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'trimestre':
        date_debut = (today - timedelta(days=89)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'annee':
        date_debut = today.replace(month=1, day=1).isoformat()
        date_fin = today.isoformat()
        
    if nature:
        sinistres = sinistres.filter(nature=nature)
    if recherche:
        sinistres = sinistres.filter(
            Q(numero_sinistre__icontains=recherche) | Q(assure__assure__numero_police__icontains=recherche)
        )
    if date_debut:
        debut_dt = datetime.combine(date.fromisoformat(date_debut), time.min)
        sinistres = sinistres.filter(date_survenance__gte=timezone.make_aware(debut_dt))
    if date_fin:
        fin_dt = datetime.combine(date.fromisoformat(date_fin), time.max)
        sinistres = sinistres.filter(date_survenance__lte=timezone.make_aware(fin_dt))

    context = {
        'chef': chef,
        'sinistres': sinistres,
        'nature_choices': Sinistre.NATURE_CHOICES,
        'nature_selectionnee': nature,
        'recherche': recherche,
        'periode_option': periode_option,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    return render(request, 'dossiers_a_valider.html', context)


# Fonction affichant au chef tous les dossiers de sinistres, tous statuts confondus
@login_required
def tous_sinistres_chef(request):
    chef = getattr(request.user, 'chef', None)
    if not chef:
        return redirect('accueil_assure')

    sinistres = Sinistre.objects.all().order_by('-date_declaration')

    statut = request.GET.get('statut', '')
    nature = request.GET.get('nature', '')
    recherche = request.GET.get('q', '').strip()
    periode_option = request.GET.get('periode_option', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')

    today = timezone.localdate()
    if periode_option == 'auj':
        date_debut = date_fin = today.isoformat()
    elif periode_option == '7j':
        date_debut = (today - timedelta(days=6)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'mois':
        date_debut = today.replace(day=1).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'trimestre':
        date_debut = (today - timedelta(days=89)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'annee':
        date_debut = today.replace(month=1, day=1).isoformat()
        date_fin = today.isoformat()

    if statut:
        sinistres = sinistres.filter(statut=statut)
    if nature:
        sinistres = sinistres.filter(nature=nature)
    if recherche:
        sinistres = sinistres.filter(
            Q(numero_sinistre__icontains=recherche) | Q(assure__assure__numero_police__icontains=recherche)
        )
    if date_debut:
        debut_dt = datetime.combine(date.fromisoformat(date_debut), time.min)
        sinistres = sinistres.filter(date_survenance__gte=timezone.make_aware(debut_dt))
    if date_fin:
        fin_dt = datetime.combine(date.fromisoformat(date_fin), time.max)
        sinistres = sinistres.filter(date_survenance__lte=timezone.make_aware(fin_dt))

    context = {
        'chef': chef,
        'sinistres': sinistres,
        'statut_choices': Sinistre.STATUS_CHOICES,
        'nature_choices': Sinistre.NATURE_CHOICES,
        'statut_selectionne': statut,
        'nature_selectionnee': nature,
        'recherche': recherche,
        'periode_option': periode_option,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    return render(request, 'tous_sinistres_chef.html', context)

   
# Fonction affichant les dossiers au statut à corriger au chef
@login_required
def dossiers_a_corriger_chef(request):
    chef = getattr(request.user, 'chef', None)
    if not chef:
        return redirect('accueil_assure')
    sinistres_a_corriger = Sinistre.objects.filter(
        statut="A_CORRIGER",
    ).order_by('date_declaration')

    nature = request.GET.get('nature', '')
    recherche = request.GET.get('q', '').strip()
    periode_option = request.GET.get('periode_option', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')

    today = timezone.localdate()
    if periode_option == 'auj':
        date_debut = date_fin = today.isoformat()
    elif periode_option == '7j':
        date_debut = (today - timedelta(days=6)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'mois':
        date_debut = today.replace(day=1).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'trimestre':
        date_debut = (today - timedelta(days=89)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'annee':
        date_debut = today.replace(month=1, day=1).isoformat()
        date_fin = today.isoformat()

    if nature:
        sinistres_a_corriger = sinistres_a_corriger.filter(nature=nature)
    if recherche:
        sinistres_a_corriger = sinistres_a_corriger.filter(
            Q(numero_sinistre__icontains=recherche) | Q(assure__assure__numero_police__icontains=recherche)
        )
    if date_debut:
        debut_dt = datetime.combine(date.fromisoformat(date_debut), time.min)
        sinistres_a_corriger = sinistres_a_corriger.filter(date_survenance__gte=timezone.make_aware(debut_dt))
    if date_fin:
        fin_dt = datetime.combine(date.fromisoformat(date_fin), time.max)
        sinistres_a_corriger = sinistres_a_corriger.filter(date_survenance__lte=timezone.make_aware(fin_dt))

    context = {
        'chef': chef,
        'sinistres_a_corriger': sinistres_a_corriger,
        'nature_choices': Sinistre.NATURE_CHOICES,
        'nature_selectionnee': nature,
        'recherche': recherche,
        'periode_option': periode_option,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    return render(request, 'dossiers_a_corriger_chef.html', context)


# Fonction affichant les dossiers au statut soumis au chef
@login_required
def dossiers_soumis_chef(request):
    chef = getattr(request.user, 'chef', None)
    if not chef:
        return redirect('accueil_assure')

    sinistres = Sinistre.objects.filter(
        statut="SOUMIS",
    ).order_by('-date_declaration')

    recherche = request.GET.get('q', '').strip()
    periode_option = request.GET.get('periode_option', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')

    today = timezone.localdate()
    if periode_option == 'auj':
        date_debut = date_fin = today.isoformat()
    elif periode_option == '7j':
        date_debut = (today - timedelta(days=6)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'mois':
        date_debut = today.replace(day=1).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'trimestre':
        date_debut = (today - timedelta(days=89)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'annee':
        date_debut = today.replace(month=1, day=1).isoformat()
        date_fin = today.isoformat()

    if recherche:
        sinistres = sinistres.filter(
            Q(assure__assure__numero_police__icontains=recherche)
        )
    if date_debut:
        debut_dt = datetime.combine(date.fromisoformat(date_debut), time.min)
        sinistres = sinistres.filter(date_survenance__gte=timezone.make_aware(debut_dt))
    if date_fin:
        fin_dt = datetime.combine(date.fromisoformat(date_fin), time.max)
        sinistres = sinistres.filter(date_survenance__lte=timezone.make_aware(fin_dt))

    context = {
        'chef': chef,
        'sinistres': sinistres,
        'recherche': recherche,
        'periode_option': periode_option,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    return render(request, 'dossiers_soumis_chef.html', context)


# Fonction affichant les détails ou informations des sinistres au chef
@login_required
def detail_sinistre_chef(request, sinistre_id):
    chef = getattr(request.user, 'chef', None)
    if not chef and not request.user.is_staff:
        return redirect('accueil_assure')

    sinistre = get_object_or_404(Sinistre, id=sinistre_id)
    retour_url = get_retour_url(request, 'dossiers_en_cours_chef')
    nb_non_lus = sinistre.messages.exclude(auteur=request.user).filter(lu=False).count()
    
    if request.method == 'POST' and 'contenu' in request.POST:
        Message.objects.create(sinistre=sinistre, auteur=request.user, contenu=request.POST.get('contenu'))
        url = reverse('detail_sinistre_chef', args=[sinistre.id])
        return redirect(f"{url}?next={retour_url}")
    
    nom_chef = request.user.get_full_name() or request.user.username
    chef_est_traitant = sinistre.agent_traitant == nom_chef
    
    sinistre.messages.exclude(auteur=request.user).filter(lu=False).update(lu=True)

    context = {
        'chef': chef,
        'sinistre': sinistre,
        'chef_est_traitant': chef_est_traitant,
        'retour_url': retour_url,
        'historique': sinistre.historique.all().order_by('date_changement'),
        'discussion': sinistre.messages.all().order_by('date_envoi'),
        'documents': sinistre.pieces.all(),
        'nb_non_lus': nb_non_lus,
    }
    return render(request, 'detail_sinistre_chef.html', context)


# Fonction permettant au chef de valider une déclaration conduisant à la génération automatique de l'attestation
@login_required
def valider_declaration(request, sinistre_id):
    chef = getattr(request.user, 'chef', None)
    if not chef and not request.user.is_staff:
        return redirect('accueil_chef')

    sinistre = get_object_or_404(Sinistre, id=sinistre_id, statut='ATTENTE_VALIDATION')
    sinistre.statut = 'EN_COURS'
    sinistre.attestation_generee = True
    sinistre.date_attestation = timezone.now()
    sinistre.chef_validateur=request.user
    sinistre.save()

    HistoriqueSinistre.objects.create(
        sinistre=sinistre,
        statut='EN_COURS',
        commentaires="Déclaration validée par le Chef de département. Attestation générée et envoyée à l'assuré.",
        auteur=request.user,
    )
    messages.success(request, f"Dossier {sinistre.numero_sinistre} validé, attestation générée.")
    return redirect(get_url_detail_dossier(request.user), sinistre_id=sinistre.id)


# Procédure de retransmission d'un dossier sinistre à l'agent
@login_required
def renvoyer_a_agent(request, sinistre_id):
    chef = getattr(request.user, 'chef', None)
    if not chef and not request.user.is_staff:
        return redirect('accueil_assure')

    # Autorise le renvoi depuis les statuts 'ATTENTE_VALIDATION' et 'EN_COURS'
    sinistre = get_object_or_404(Sinistre, id=sinistre_id, statut__in=['ATTENTE_VALIDATION', 'EN_COURS'])

    if request.method == 'POST':
        commentaires = request.POST.get('commentaires') or request.POST.get('motif') or "Demander la révision."
        statut_precedent = sinistre.statut

        sinistre.statut = 'A_CORRIGER'
        sinistre.indemnisation_validee = False
        sinistre.save()

        if statut_precedent == 'EN_COURS':
            commentaire_historique = f"Prix à réviser - {commentaires}"
        else:
            commentaire_historique = commentaires

        HistoriqueSinistre.objects.create(
            sinistre=sinistre,
            statut='A_CORRIGER',
            commentaires=commentaire_historique,
            auteur=request.user,
        )

        nom_chef = request.user.get_full_name() or request.user.username
        chef_est_traitant = sinistre.agent_traitant == nom_chef
        
        if chef_est_traitant:
            messages.warning(
                request, 
                f"Le dossier {sinistre.numero_sinistre} a été repassé en correction."
                f"Comme vous êtes vous-même en charge de ce dossier, vous pouvez modifier le prix retenu directement ci-dessous."
                )
        else:
            messages.warning(
                request,
                f"Le dossier {sinistre.numero_sinistre} a été renvoyé au rédacteur ({sinistre.agent_traitant}) pour révision."
            )
            
        return redirect(get_url_detail_dossier(request.user), sinistre_id=sinistre.id)

    return redirect(get_url_detail_dossier(request.user), sinistre_id=sinistre.id)


# Procédure permettant au chef de  valider une indemnisation saisie par l'assuré
@login_required
def valider_indemnisation(request, sinistre_id):
    chef = getattr(request.user, 'chef', None)
    if not chef and not request.user.is_staff:
        return redirect('accueil_assure')

    sinistre = get_object_or_404(Sinistre, id=sinistre_id, statut__in=['EN_COURS','ATTENTE_VALIDATION','A_CORRIGER'])

    if sinistre.prix_retenu is None:
        messages.error(request, "L'agent doit d'abord saisir le prix retenu.")
        return redirect(get_url_detail_dossier(request.user), sinistre_id=sinistre.id)

    sinistre.indemnisation_validee = True
    sinistre.statut = 'EN_COURS'
    sinistre.save()
    if sinistre.prix_retenu == 0:
        commentaire = "Sinistre non indemnisable (prix retenu fixé à 0 FCFA), validé par le chef de département."
    else:
        commentaire = f"Indemnisation de {sinistre.prix_retenu} FCFA validée par le chef de département."
        
    HistoriqueSinistre.objects.create(
        sinistre=sinistre,
        statut="EN_COURS",
        commentaires=commentaire,
        auteur=request.user,
    )
    messages.success(request, "Indemnisation validée. Le dossier peut maintenant être clôturé ou classé sans suite.")
    return redirect(get_url_detail_dossier(request.user), sinistre_id=sinistre.id)


# Fonction permettant au chef de demander la révision du prix retenu
@login_required
def demander_revision_prix(request, sinistre_id):
    chef = getattr(request.user, 'chef', None)
    if not chef and not request.user.is_staff:
        return redirect('accueil_assure')

    # Récupération du sinistre en attente de validation ou en cours
    sinistre = get_object_or_404(Sinistre, id=sinistre_id, statut__in=['ATTENTE_VALIDATION', 'EN_COURS'])

    if request.method == 'POST':
        commentaires = request.POST.get('commentaires', '').strip()

        # 1. Passer le dossier au statut 'A_CORRIGER'
        sinistre.statut = 'A_CORRIGER'
        sinistre.indemnisation_validee = False
        sinistre.save()

        # 2. Historiser le motif explicatif
        HistoriqueSinistre.objects.create(
            sinistre=sinistre,
            statut='A_CORRIGER',
            commentaires=commentaires or "Demande de révision envoyée à l'agent.",
            auteur=request.user,
        )

        # 3. Notification de succès pour le chef
        messages.success(
            request, 
            f"✅ Le dossier {sinistre.numero_sinistre} a bien été renvoyé dans la file d'attente de l'agent."
        )

        # Redirection vers la liste des dossiers à valider du chef
        return redirect('dossiers_a_valider') 

    return render(request, 'demander_revision.html', {'sinistre': sinistre})


# Procédure permettant à l'agent de cloturé un sinistre
@login_required
def clore_sinistre(request, sinistre_id):
    chef = getattr(request.user, 'chef', None)
    if not chef and not request.user.is_staff:
        return redirect('accueil_assure')

    sinistre = get_object_or_404(Sinistre, id=sinistre_id, statut='EN_COURS')
    sinistre.statut = 'CLOTURE'
    sinistre.date_cloture = timezone.now()
    sinistre.save()
    HistoriqueSinistre.objects.create(
        sinistre=sinistre,
        statut='CLOTURE',
        commentaires="Flux financiers constatés, dossier clôturé par le Chef de département.",
        auteur=request.user,
    )
    messages.success(request, f"Dossier {sinistre.numero_sinistre} clôturé.")
    return redirect(get_url_detail_dossier(request.user), sinistre_id=sinistre.id)


# Procédure permettant à l'agent de classer un sinistre sans suite
@login_required
def classer_sans_suite(request, sinistre_id):
    chef = getattr(request.user, 'chef', None)
    if not chef and not request.user.is_staff:
        return redirect('accueil_assure')

    sinistre = get_object_or_404(Sinistre, id=sinistre_id, statut='EN_COURS')

    if request.method == 'POST':
        form = SansSuiteForm(request.POST)
        if form.is_valid():
            sinistre.statut = 'SANS_SUITE'
            sinistre.motif_sans_suite = form.cleaned_data['motif']
            sinistre.save()
            HistoriqueSinistre.objects.create(
                sinistre=sinistre,
                statut='SANS_SUITE',
                commentaires=form.cleaned_data['motif'],
                auteur=request.user,
            )
            messages.success(request, f"Dossier {sinistre.numero_sinistre} classé sans suite.")
            return redirect(get_url_detail_dossier(request.user), sinistre_id=sinistre.id)
    else:
        form = SansSuiteForm()

    return render(request, 'classer_sans_suite.html', {'sinistre': sinistre, 'form': form})


# Procédure permettant à l'agent de réouvrir un sinistre déjà clôturé ou classer sans suite
@login_required
def reouvrir_dossier(request, sinistre_id):
    chef = getattr(request.user, 'chef', None)
    if not chef and not request.user.is_staff:
        return redirect('accueil_assure')

    sinistre = get_object_or_404(Sinistre, id=sinistre_id, statut__in=['CLOTURE', 'SANS_SUITE'])
    sinistre.statut = 'REOUVERT'
    sinistre.save()
    HistoriqueSinistre.objects.create(
        sinistre=sinistre,
        statut='REOUVERT',
        commentaires="Dossier réouvert par le Chef de département pour ré-instruction.",
        auteur=request.user,
    )
    messages.success(request, f"Dossier {sinistre.numero_sinistre} réouvert.")
    return redirect(get_url_detail_dossier(request.user), sinistre_id=sinistre.id)


# Fonction affichant au chef les sinistres en cours
@login_required
def dossiers_en_cours_chef(request):
    chef = getattr(request.user, 'chef', None)
    if not chef:
        return redirect('accueil_assure')

    sinistres = Sinistre.objects.filter(
        statut__in=['ATTENTE_VALIDATION', 'EN_COURS', 'A_CORRIGER', 'REOUVERT'],
    ).order_by('-date_declaration')

    nature = request.GET.get('nature', '')
    recherche = request.GET.get('q', '').strip()
    periode_option = request.GET.get('periode_option', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')

    today = timezone.localdate()
    if periode_option == 'auj':
        date_debut = date_fin = today.isoformat()
    elif periode_option == '7j':
        date_debut = (today - timedelta(days=6)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'mois':
        date_debut = today.replace(day=1).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'trimestre':
        date_debut = (today - timedelta(days=89)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'annee':
        date_debut = today.replace(month=1, day=1).isoformat()
        date_fin = today.isoformat()

    if nature:
        sinistres = sinistres.filter(nature=nature)
    if recherche:
        sinistres = sinistres.filter(
            Q(numero_sinistre__icontains=recherche) | Q(assure__assure__numero_police__icontains=recherche)
        )
    if date_debut:
        debut_dt = datetime.combine(date.fromisoformat(date_debut), time.min)
        sinistres = sinistres.filter(date_survenance__gte=timezone.make_aware(debut_dt))
    if date_fin:
        fin_dt = datetime.combine(date.fromisoformat(date_fin), time.max)
        sinistres = sinistres.filter(date_survenance__lte=timezone.make_aware(fin_dt))

    context = {
        'chef': chef,
        'sinistres': sinistres,
        'nature_choices': Sinistre.NATURE_CHOICES,
        'nature_selectionnee': nature,
        'recherche': recherche,
        'periode_option': periode_option,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    return render(request, 'dossiers_en_cours_chef.html', context)
 

# Fonction affichant au chef les sinistres clôturés
@login_required
def dossiers_clotures_chef(request):
    chef = getattr(request.user, 'chef', None)
    if not chef:
        return redirect('accueil_assure')

    sinistres = Sinistre.objects.filter(
        statut__in=['CLOTURE', 'SANS_SUITE'],
    ).order_by('-date_declaration')

    nature = request.GET.get('nature', '')
    recherche = request.GET.get('q', '').strip()
    periode_option = request.GET.get('periode_option', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')

    today = timezone.localdate()
    if periode_option == 'auj':
        date_debut = date_fin = today.isoformat()
    elif periode_option == '7j':
        date_debut = (today - timedelta(days=6)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'mois':
        date_debut = today.replace(day=1).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'trimestre':
        date_debut = (today - timedelta(days=89)).isoformat()
        date_fin = today.isoformat()
    elif periode_option == 'annee':
        date_debut = today.replace(month=1, day=1).isoformat()
        date_fin = today.isoformat()

    if nature:
        sinistres = sinistres.filter(nature=nature)
    if recherche:
        sinistres = sinistres.filter(
            Q(numero_sinistre__icontains=recherche) | Q(assure__assure__numero_police__icontains=recherche)
        )
    if date_debut:
        debut_dt = datetime.combine(date.fromisoformat(date_debut), time.min)
        sinistres = sinistres.filter(date_survenance__gte=timezone.make_aware(debut_dt))
    if date_fin:
        fin_dt = datetime.combine(date.fromisoformat(date_fin), time.max)
        sinistres = sinistres.filter(date_survenance__lte=timezone.make_aware(fin_dt))

    context = {
        'chef': chef,
        'sinistres': sinistres,
        'nature_choices': Sinistre.NATURE_CHOICES,
        'nature_selectionnee': nature,
        'recherche': recherche,
        'periode_option': periode_option,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    return render(request, 'dossiers_clotures_chef.html', context)


# Fonction permettant de modifier son mot de passe
@login_required
def changer_mot_de_passe_chef(request):
    chef = getattr(request.user, 'chef', None)
    if not chef:
        return redirect('accueil_assure')

    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, form.user)
            chef.doit_changer_mot_de_passe = False
            chef.save()
            messages.success(request, "Mot de passe modifié avec succès.")
            return redirect('accueil_chef')
    else:
        form = PasswordChangeForm(request.user)

    return render(request, 'changer_mot_de_passe_chef.html', {'form': form})


# Fonction affichant le profil du chef
@login_required
def profil_chef(request):
    chef = getattr(request.user, 'chef', None)
    if not chef:
        return redirect('accueil_assure')
    return render(request, 'profil_chef.html', {'chef': chef})


# Fonction permettant au chef de modifier son profil
@login_required
def modifier_profil_chef(request):
    chef = getattr(request.user, 'chef', None)
    if not chef:
        return redirect('accueil_assure')
    
    if request.method == 'POST':
        form = ModifierProfilForm(request.POST)
        if form.is_valid():
            request.user.email = form.cleaned_data['email']
            request.user.save(update_fields=['email'])
            chef.telephone = form.cleaned_data['telephone']
            chef.save(update_fields=['telephone'])
            messages.success(request, "Vos informations on été mis à jour avec succès.")
            return redirect('profil_chef')
    else:
        form = ModifierProfilForm(initial={'email':request.user.email, 'telephone':chef.telephone})
    return render(request, 'modifier_profil_chef.html', {'form':form})


#-------------------------------------------------------------------
#                   ATTESTATIONS
#-------------------------------------------------------------------


# Fonction permettant d'ouvrir et de voir une attestation
@login_required
def voir_attestation(request, sinistre_id):
    sinistre = get_object_or_404(
        Sinistre, 
        id=sinistre_id, 
    )
    is_owner = (getattr(sinistre, 'assure', None) == request.user or getattr(sinistre, 'assure_id', None) == request.user.id)
    
    if is_owner or hasattr(request.user, 'agent') or hasattr(request.user, 'chef') or request.user.is_staff:
        quittances = sinistre.quittances.all().order_by('date_debut')
        date_effet = quittances.first().date_debut if quittances.exists() else None
        date_echeance = quittances.first().date_fin if quittances.exists() else None
        return render(request, 'attestation.html', {
            'sinistre': sinistre,
            'date_effet': date_effet,
            'date_echeance': date_echeance,
        })
        
    return redirect('accueil_assure')


# Fonction permettant de télécharger une attestation
def link_callback_pdf(uri, rel):
    """Résout les chemins static/media en chemins de fichiers locaux pour xhtml2pdf."""
    if uri.startswith(settings.STATIC_URL):
        path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
        if not os.path.isfile(path):
            for static_dir in settings.STATICFILES_DIRS:
                candidate = os.path.join(static_dir, uri.replace(settings.STATIC_URL, ""))
                if os.path.isfile(candidate):
                    path = candidate
                    break
        return path
    elif uri.startswith(settings.MEDIA_URL):
        return os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))
    return uri


@login_required
def telecharger_attestation(request, sinistre_id):
    sinistre = get_object_or_404(Sinistre, id=sinistre_id, attestation_generee=True)
    is_owner = (getattr(sinistre, 'assure', None) == request.user or getattr(sinistre, 'assure_id', None) == request.user.id)

    if not (is_owner or hasattr(request.user, 'agent') or hasattr(request.user, 'chef')):
        return redirect('accueil_assure')

    quittances = sinistre.quittances.all().order_by('date_debut')
    date_effet = quittances.first().date_debut if quittances.exists() else None
    date_echeance = quittances.first().date_fin if quittances.exists() else None

    logo_path = os.path.join(settings.STATICFILES_DIRS[0], "Logo Fidelia.jpeg")
    brand_path = os.path.join(settings.STATICFILES_DIRS[0], "images", "brand_scriptbold.png")
    
    signature_path = ""
    if sinistre.chef_validateur and getattr(sinistre.chef_validateur.chef, 'signature', None):
        signature_path = sinistre.chef_validateur.chef.signature.path

    html_string = render_to_string('attestation_pdf.html', {
        'sinistre': sinistre,
        'date_effet': date_effet,
        'date_echeance': date_echeance,
        'logo_path': logo_path,
        'brand_path': brand_path,
        'signature_path': signature_path,
    }, request=request)

    result = io.BytesIO()
    pisa_status = pisa.CreatePDF(html_string, dest=result, link_callback=link_callback_pdf)

    if pisa_status.err:
        return HttpResponse('Erreur lors de la génération du PDF', status=500)

    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="attestation_{sinistre.numero_sinistre}.pdf"'
    return response


#-------------------------------------------------------------------
#                   ESPACE ADMINISTRATEUR
#-------------------------------------------------------------------


# Page d'acceuil de l'administrateur
@login_required
@user_passes_test(lambda u: u.is_staff)
def accueil_admin(request):
    sinistres = Sinistre.objects.all().order_by('-date_survenance')
    debut_mois = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    today = timezone.localdate()
    debut_jour = timezone.make_aware(datetime.combine(today, time.min))
    fin_jour = timezone.make_aware(datetime.combine(today, time.max))
    
    sinistres_du_jour = sinistres.filter(
        date_declaration__gte=debut_jour,
        date_declaration__lte=fin_jour,
    ).order_by('-date_declaration')
    
    
    context = {
        'nb_sinistres_total': sinistres.count(),
        'nb_sinistres_en_cours': sinistres.exclude(statut__in=['CLOTURE', 'SANS_SUITE']).count(),
        'nb_sinistres_clotures': sinistres.filter(statut='CLOTURE').count(),
        'nb_sinistres_mois': sinistres.filter(date_declaration__gte=debut_mois).count(),
        'nb_agences': Agence.objects.count(),
        'nb_agents': Agent.objects.count(),
        'nb_chefs': ChefDepartement.objects.count(),
        'nb_assures': Assure.objects.count(),
        'derniers_sinistres': sinistres_du_jour,
    }
    return render(request, 'accueil_admin.html', context)


# Génère le prochain matricule disponible pour un préfixe donné (AG ou CH), incrément global sur 4 chiffres
def generer_prochain_matricule(model, prefixe):
    dernier = model.objects.filter(matricule__startswith=f"{prefixe}-").order_by('-matricule').first()
    if dernier:
        try:
            dernier_numero = int(dernier.matricule.split('-')[-1])
        except ValueError:
            dernier_numero = 0
    else:
        dernier_numero = 0
    nouveau_numero = dernier_numero + 1
    return f"{prefixe}-{nouveau_numero:04d}"


# Fonction permettant à l'administrateur de créer un agent
@login_required
@user_passes_test(lambda u: u.is_staff)
def creer_agent(request):
    if request.method == 'POST':
        form = AgentCreationForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            username = f"{data['prenom']}.{data['nom']}".lower().replace(' ', '')

            user = User.objects.create_user(
                username=username,
                email=data['email'],
                first_name=data['prenom'],
                last_name=data['nom'],
            )
            user.set_password('0000')
            user.save()

            matricule = generer_prochain_matricule(Agent, 'AG')
            
            Agent.objects.create(
                user=user,
                agence=data['agence'],
                matricule=matricule,
                telephone=data['telephone'],
                compte_active=False,
                doit_changer_mot_de_passe=True,
            )
            messages.success(request, f"Rédacteur créé. Identifiant : {username} — Matricule : {matricule} - Mot de passe temporaire : 0000")
            return redirect('creer_agent')
    else:
        form = AgentCreationForm()

    return render(request, 'creer_agent.html', {'form': form})


# Fonction permettant à l'administrateur d'accéder à la liste des agents
@login_required
@user_passes_test(lambda u: u.is_staff)
def liste_agents(request):
    agents = Agent.objects.select_related('user', 'agence').order_by('user__last_name')

    recherche = request.GET.get('q', '').strip()
    statut = request.GET.get('statut', '')
    agence_id = request.GET.get('agence', '')

    if recherche:
        agents = agents.filter(
            Q(user__last_name__icontains=recherche) |
            Q(user__first_name__icontains=recherche) |
            Q(matricule__icontains=recherche) |
            Q(user__email__icontains=recherche)
        )
    if statut == 'actif':
        agents = agents.filter(user__is_active=True)
    elif statut == 'inactif':
        agents = agents.filter(user__is_active=False)
    if agence_id:
        agents = agents.filter(agence_id=agence_id)

    context = {
        'agents': agents,
        'recherche': recherche,
        'statut_selectionne': statut,
        'agence_selectionnee': agence_id,
        'agences': Agence.objects.order_by('nom'),
    }
    return render(request, 'liste_agents.html', context)


# Fonction permettant à l'administrateur d'agir sur plusieurs rédacteurs sélectionnés (activer/désactiver/supprimer)
@login_required
@user_passes_test(lambda u: u.is_staff)
def action_lot_agents(request):
    if request.method == 'POST':
        ids = request.POST.getlist('selection')
        action = request.POST.get('action')
        agents = Agent.objects.filter(id__in=ids)
        nb = agents.count()

        if nb == 0:
            messages.warning(request, "Aucun rédacteur sélectionné.")
        elif action == 'activer':
            for agent in agents:
                agent.user.is_active = True
                agent.user.save()
            messages.success(request, f"{nb} rédacteur(s) activé(s).")
        elif action == 'desactiver':
            for agent in agents:
                agent.user.is_active = False
                agent.user.save()
            messages.success(request, f"{nb} rédacteur(s) désactivé(s).")
        elif action == 'supprimer':
            for agent in agents:
                agent.user.delete()
            messages.success(request, f"{nb} rédacteur(s) supprimé(s).")
        else:
            messages.error(request, "Action inconnue.")

    return redirect('liste_agents')


# Fonction permettant à l'administrateur d'agir sur plusieurs rédacteurs sélectionnés (activer/désactiver/supprimer)
@login_required
@user_passes_test(lambda u: u.is_staff)
def action_lot_chefs(request):
    if request.method == 'POST':
        ids = request.POST.getlist('selection')
        action = request.POST.get('action')
        chefs = ChefDepartement.objects.filter(id__in=ids)
        nb = chefs.count()

        if nb == 0:
            messages.warning(request, "Aucun chef sélectionné.")
        elif action == 'activer':
            for chef in chefs:
                chef.user.is_active = True
                chef.user.save()
            messages.success(request, f"{nb} chef(s) activé(s).")
        elif action == 'desactiver':
            for chef in chefs:
                chef.user.is_active = False
                chef.user.save()
            messages.success(request, f"{nb} chef(s) désactivé(s).")
        elif action == 'supprimer':
            for chef in chefs:
                chef.user.delete()
            messages.success(request, f"{nb} chef(s) supprimé(s).")
        else:
            messages.error(request, "Action inconnue.")

    return redirect('liste_chefs')


# Fonction permettant à l'administrateur de modifier les informations d'un assure
@login_required
@user_passes_test(lambda u: u.is_staff)
def modifier_assure_admin(request, assure_id):
    assure = get_object_or_404(Assure, id=assure_id)
    if request.method == 'POST':
        form = AssureAdminForm(request.POST, instance=assure)
        if form.is_valid():
            form.save()
            messages.success(request, "Assuré mis à jour.")
            return redirect('liste_assures')
    else:
        form = AssureAdminForm(instance=assure)

    return render(request, 'modifier_assure_admin.html', {
        'form': form, 'assure': assure,
    })
        
    
# Fonction permettant à l'administrateur d'agir sur plusieurs assurés sélectionnés (activer/désactiver/supprimer)
@login_required
@user_passes_test(lambda u: u.is_staff)
def action_lot_assures(request):
    if request.method == 'POST':
        ids = request.POST.getlist('selection')
        action = request.POST.get('action')
        assures = Assure.objects.filter(id__in=ids)
        nb = assures.count()

        if nb == 0:
            messages.warning(request, "Aucun assuré sélectionné.")
        elif action == 'activer':
            for assure in assures:
                assure.user.is_active = True
                assure.user.save()
            messages.success(request, f"{nb} assuré(s) activé(s).")
        elif action == 'desactiver':
            for assure in assures:
                assure.user.is_active = False
                assure.user.save()
            messages.success(request, f"{nb} assuré(s) désactivé(s).")
        elif action == 'supprimer':
            for assure in assures:
                assure.user.delete()
            messages.success(request, f"{nb} assuré(s) supprimé(s).")
        else:
            messages.error(request, "Action inconnue.")

    return redirect('liste_assures')


# Fonction permettant à l'administrateur de supprimer définitivement le compte d'un assuré
@login_required
@user_passes_test(lambda u: u.is_staff)
def supprimer_assure_admin(request, assure_id):
    assure = get_object_or_404(Assure, id=assure_id)
    if request.method == 'POST':
        nom_complet = assure.user.get_full_name() or assure.user.username
        assure.user.delete()
        messages.success(request, f"Le compte de {nom_complet} a été supprimé.")
        return redirect('liste_assures')

    return render(request, 'confirmer_supression.html', {
        'objet_nom': assure.user.get_full_name() or assure.user.username,
        'type_objet': 'assuré',
        'annuler_url': 'liste_assures',
        'confirmer_url': 'supprimer_assure_admin',
        'objet_id': assure.id,
    })
    
    
# Fonction permettant à l'administrateur de modifier les informations d'un agent
@login_required
@user_passes_test(lambda u: u.is_staff)
def modifier_agent_admin(request, agent_id):
    agent = get_object_or_404(Agent, id=agent_id)
    if request.method == 'POST':
        user_form = ModifierProfilAdminForm(request.POST, instance=agent.user)
        agent_form = ModifierAgentAdminForm(request.POST, instance=agent)
        if user_form.is_valid() and agent_form.is_valid():
            user_form.save()
            agent_form.save()
            messages.success(request, "Agent mis à jour.")
            return redirect('liste_agents')
    else:
        user_form = ModifierProfilAdminForm(instance=agent.user)
        agent_form = ModifierAgentAdminForm(instance=agent)
    return render(request, 'modifier_agent_admin.html', {
        'user_form': user_form, 'agent_form': agent_form, 'agent': agent,
    })


# Fonction permettant à l'administrateur d'activer ou de désactiver le compte d'un agent
@login_required
@user_passes_test(lambda u: u.is_staff)
def toggle_agent_actif(request, agent_id):
    agent = get_object_or_404(Agent, id=agent_id)
    agent.user.is_active = not agent.user.is_active
    agent.user.save()
    messages.success(request, f"Compte {'réactivé' if agent.user.is_active else 'désactivé'}.")
    return redirect('liste_agents')


# Fonction permettant à l'administrateur de réactiver le compte d'un agent
@login_required
@user_passes_test(lambda u: u.is_staff)
def reinitialiser_mdp_agent(request, agent_id):
    agent = get_object_or_404(Agent, id=agent_id)
    agent.user.set_password('0000')
    agent.user.save()
    agent.doit_changer_mot_de_passe = True
    agent.save()
    messages.success(request, f"Mot de passe réinitialisé à 0000 pour {agent.user.get_full_name() or agent.user.username}.")
    return redirect('liste_agents')


# Fonction permettant à l'administrateur de supprimer le compte d'un agent
@login_required
@user_passes_test(lambda u: u.is_staff)
def supprimer_agent(request, agent_id):
    agent = get_object_or_404(Agent, id=agent_id)
    if request.method == 'POST':
        agent.user.delete()
        messages.success(request, "Agent supprimé.")
        return redirect('liste_agents')
    return render(request, 'confirmer_suppression.html', {
        'objet_nom': agent.user.get_full_name() or agent.user.username,
        'type_objet': 'agent',
        'annuler_url': 'liste_agents',
        'confirmer_url': 'supprimer_agent',
        'objet_id': agent.id,
    })


# Fonction permettant à l'administrateur de créer le compte d'un chef
@login_required
@user_passes_test(lambda u: u.is_staff)
def creer_chef(request):
    if request.method == 'POST':
        form = ChefCreationForm(request.POST, request.FILES)
        if form.is_valid():
            data = form.cleaned_data
            username = f"{data['prenom']}.{data['nom']}".lower().replace(' ', '')

            user = User.objects.create_user(
                username=username,
                email=data['email'],
                first_name=data['prenom'],
                last_name=data['nom'],
            )
            user.set_password('0000')
            user.save()

            matricule = generer_prochain_matricule(ChefDepartement, 'CH')
            
            ChefDepartement.objects.create(
                user=user,
                agence=data['agence'],
                matricule=matricule,
                telephone=data['telephone'],
                signature=data['signature'],
                doit_changer_mot_de_passe=True,
            )
            messages.success(request, f"Chef créé. Identifiant : {username} — Matricule : {matricule} - Mot de passe temporaire : 0000")
            return redirect('creer_chef')
    else:
        form = ChefCreationForm()

    return render(request, 'creer_chef.html', {'form': form})


# Fonction affichant à l'administrateur la liste des chefs
@login_required
@user_passes_test(lambda u: u.is_staff)
def liste_chefs(request):
    chefs = ChefDepartement.objects.select_related('user', 'agence').order_by('user__last_name')

    recherche = request.GET.get('q', '').strip()
    statut = request.GET.get('statut', '')
    agence_id = request.GET.get('agence', '')

    if recherche:
        chefs = chefs.filter(
            Q(user__last_name__icontains=recherche) |
            Q(user__first_name__icontains=recherche) |
            Q(matricule__icontains=recherche) |
            Q(user__email__icontains=recherche)
        )
    if statut == 'actif':
        chefs = chefs.filter(user__is_active=True)
    elif statut == 'inactif':
        chefs = chefs.filter(user__is_active=False)
    if agence_id:
        chefs = chefs.filter(agence_id=agence_id)

    context = {
        'chefs': chefs,
        'recherche': recherche,
        'statut_selectionne': statut,
        'agence_selectionnee': agence_id,
        'agences': Agence.objects.order_by('nom'),
    }
    return render(request, 'liste_chefs.html', context)


# Fonction permettant à l'administrateur de modifier les informations d'un chef
@login_required
@user_passes_test(lambda u: u.is_staff)
def modifier_chef_admin(request, chef_id):
    chef = get_object_or_404(ChefDepartement, id=chef_id)
    if request.method == 'POST':
        user_form = ModifierProfilAdminForm(request.POST, instance=chef.user)
        chef_form = ModifierChefAdminForm(request.POST, instance=chef)
        if user_form.is_valid() and chef_form.is_valid():
            user_form.save()
            chef_form.save()
            messages.success(request, "Chef mis à jour.")
            return redirect('liste_chefs')
    else:
        user_form = ModifierProfilAdminForm(instance=chef.user)
        chef_form = ModifierChefAdminForm(instance=chef)
    return render(request, 'modifier_chef_admin.html', {
        'user_form': user_form, 'chef_form': chef_form, 'chef': chef,
    })


# Fonction permettant à l'administrateur d'activer ou de désactiver le compte d'un chef
@login_required
@user_passes_test(lambda u: u.is_staff)
def toggle_chef_actif(request, chef_id):
    chef = get_object_or_404(ChefDepartement, id=chef_id)
    chef.user.is_active = not chef.user.is_active
    chef.user.save()
    messages.success(request, f"Compte {'réactivé' if chef.user.is_active else 'désactivé'}.")
    return redirect('liste_chefs')


# Fonction permettant à l'administrateur de réactiver le compte d'un chef
@login_required
@user_passes_test(lambda u: u.is_staff)
def reinitialiser_mdp_chef(request, chef_id):
    chef = get_object_or_404(ChefDepartement, id=chef_id)
    chef.user.set_password('0000')
    chef.user.save()
    chef.doit_changer_mot_de_passe = True
    chef.save()
    messages.success(request, f"Mot de passe réinitialisé à 0000 pour {chef.user.get_full_name() or chef.user.username}.")
    return redirect('liste_chefs')


# Fonction permettant à l'administrateur de supprimer le compte d'un chef
@login_required
@user_passes_test(lambda u: u.is_staff)
def supprimer_chef(request, chef_id):
    chef = get_object_or_404(ChefDepartement, id=chef_id)
    if request.method == 'POST':
        chef.user.delete()
        messages.success(request, "Chef supprimé.")
        return redirect('liste_chefs')
    return render(request, 'confirmer_suppression.html', {
        'objet_nom': chef.user.get_full_name() or chef.user.username,
        'type_objet': 'chef',
        'annuler_url': 'liste_chefs',
        'confirmer_url': 'supprimer_chef',
        'objet_id': chef.id,
    })


# Fonction affichant à l'administrateur la liste des assurés
@login_required
@user_passes_test(lambda u: u.is_staff)
def liste_assures(request):
    assures = Assure.objects.select_related('user', 'agence').order_by('user__last_name')

    recherche = request.GET.get('q', '').strip()
    statut = request.GET.get('statut', '')
    activation = request.GET.get('activation', '')

    if recherche:
        assures = assures.filter(
            Q(user__last_name__icontains=recherche) |
            Q(user__first_name__icontains=recherche) |
            Q(numero_police__icontains=recherche) |
            Q(user__email__icontains=recherche)
        )
    if statut == 'actif':
        assures = assures.filter(user__is_active=True)
    elif statut == 'inactif':
        assures = assures.filter(user__is_active=False)
    if activation == 'active':
        assures = assures.filter(compte_active=True)
    elif activation == 'attente':
        assures = assures.filter(compte_active=False)

    context = {
        'assures': assures,
        'recherche': recherche,
        'statut_selectionne': statut,
        'activation_selectionnee': activation,
    }
    return render(request, 'liste_assures.html', context)


# Fonction permettant à l'administrateur d'activer ou de désactiver le compte d'un assuré
@login_required
@user_passes_test(lambda u: u.is_staff)
def toggle_assure_actif(request, assure_id):
    assure = get_object_or_404(Assure, id=assure_id)
    assure.user.is_active = not assure.user.is_active
    assure.user.save()
    messages.success(request, f"Compte {'réactivé' if assure.user.is_active else 'désactivé'}.")
    return redirect('liste_assures')


# Supervision globale de tous les sinistres pour l'administrateur.
@login_required
@user_passes_test(lambda u: u.is_staff)
def supervision_sinistres(request):
    sinistres = Sinistre.objects.select_related('assure', 'vehicule', 'region').order_by('-date_declaration')

    statut = request.GET.get('statut', '')
    nature = request.GET.get('nature', '')
    recherche = request.GET.get('q', '').strip()

    if statut:
        sinistres = sinistres.filter(statut=statut)
    if nature:
        sinistres = sinistres.filter(nature=nature)
    if recherche:
        sinistres = sinistres.filter(
            Q(numero_sinistre__icontains=recherche) | Q(assure__assure__numero_police__icontains=recherche)
        )

    context = {
        'sinistres': sinistres,
        'statut_choices': Sinistre.STATUS_CHOICES,
        'nature_choices': Sinistre.NATURE_CHOICES,
        'statut_selectionne': statut,
        'nature_selectionnee': nature,
        'recherche': recherche,
    }
    return render(request, 'supervision_sinistres.html', context)


# Fonction affichant le profil de l'administrateur
@login_required
@user_passes_test(lambda u: u.is_staff)
def profil_admin(request):
    return render(request, 'profil_admin.html')


# Fonction permettant à l'administrateur de modifier son profil
@login_required
@user_passes_test(lambda u: u.is_staff)
def modifier_profil_admin(request):
    if request.method == 'POST':
        form = ModifierProfilAdminForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Vos informations ont été mises à jour.")
            return redirect('profil_admin')
    else:
        form = ModifierProfilAdminForm(instance=request.user)
    return render(request, 'modifier_profil_admin.html', {'form': form})
   
        
# Fonction permettant à l'administrateur d'importer des contrats via un fichier Excel
@login_required
@user_passes_test(lambda u: u.is_staff)
def importer_donnees_admin(request):
    if request.method == 'POST':
        form = ImportExcelForm(request.POST, request.FILES)
        if form.is_valid():
            fichier = request.FILES['fichier']
            try:
                df = pd.read_excel(fichier)
                df.columns = [str(c).strip() for c in df.columns]
                
                def get_val(row, possible_keys):
                    for k in possible_keys:
                        for col in row.index:
                            if str(col).strip().lower().replace(' ', '_') == str(k).strip().lower().replace(' ', '_'):
                                val = row[col]
                                if pd.notna(val):
                                    return val
                    return None

                lignes_importees = 0
                for index, row in df.iterrows():
                    branche = str(get_val(row, ['BRANCHE', 'TYPE CONTRAT', 'CONTRAT']) or '').strip()
                    if branche.lower() != 'automobile':
                        continue  
                        
                    email_excel = get_val(row, ['EMAIL', 'COURRIEL', 'MAIL'])
                    if not email_excel:
                        continue
                    
                    if request.user.is_authenticated and email_excel == request.user.email:
                        continue
                    
                    if User.objects.filter(email=email_excel, is_staff=True).exists():
                        continue

                    numero_police = get_val(row, ['NUMERO_POLICE', 'NUMERO POLICE', 'NUMERO_CONTRAT', 'NUMERO CONTRAT'])
                    prenom = get_val(row, ['PRENOM', 'PRENOMS'])
                    nom = get_val(row, ['NOM'])
                    telephone = get_val(row, ['TELEPHONE', 'TEL'])
                    numero_quittance = get_val(row, ['NUMERO_QUITTANCE', 'NUMERO QUITTANCE', 'QUITTANCE'])
                    
                    if not numero_police or not numero_quittance:
                        continue

                    marque = get_val(row, ['MARQUE', 'MARQUE_VEHICULE']) or ''
                    categorie = get_val(row, ['CATEGORIE', 'CATEGORIE_VEHICULE']) or None
                    immatriculation = get_val(row, ['IMMATRICULATION', 'IMMAT'])

                    user, user_created = User.objects.get_or_create(
                        email=email_excel,
                        defaults={
                            'username': email_excel,
                            'first_name': prenom or '',
                            'last_name': nom or '',
                            'is_active': False,
                        }
                    )
                    
                    if user.username == 'admin_fidelia' or user.is_staff:
                        continue

                    if prenom:
                        user.first_name = prenom
                    if nom:
                        user.last_name = nom
                    if user_created:
                        user.is_active = False  
                    user.save()

                    assure, created = Assure.objects.get_or_create(
                        numero_police=numero_police,
                        defaults={
                            'user': user,
                            'telephone': telephone,
                            'compte_active': False,
                        }
                    )

                    if not created:
                        assure.user = user
                        if telephone:
                            assure.telephone = telephone
                        assure.save()

                    date_debut = get_val(row, ['DATE_DEBUT', 'DATE DEBUT', 'DATE_EFFET', 'DATE EFFET'])
                    date_fin = get_val(row, ['DATE_FIN', 'DATE FIN', 'DATE_ECHEANCE', 'DATE ECHEANCE'])
                    prime = get_val(row, ['PRIME_TTC', 'PRIME TTC', 'PRIME'])

                    quittance, _ = Quittance.objects.update_or_create(
                        numero_quittance=numero_quittance,
                        defaults={
                            'contrat': assure,
                            'branche': branche,
                            'date_debut': date_debut,
                            'date_fin': date_fin,
                            'prime': prime
                        }
                    )

                    if immatriculation and marque:
                        lignes_importees += 1
                        Vehicule.objects.update_or_create(
                            immatriculation=immatriculation,
                            defaults={
                                'marque': marque,
                                'categorie': categorie,
                                'proprietaire': user,
                                'quittance': quittance,
                            }
                        )
                
                if lignes_importees > 0:
                    messages.success(request, f"{lignes_importees} contrat(s) importé(s) avec succès !")
                else:
                    messages.warning(request, "Aucune ligne n'a pu être importée. Vérifiez les valeurs de TYPE_CONTRAT, EMAIL, NUMERO_POLICE et NUMERO_QUITTANCE dans votre fichier.")
                return redirect('liste_contrats')
                
            except Exception as e:
                messages.error(request, f"Erreur lors de l'importation : {e}")
    else:
        form = ImportExcelForm()
    
    return render(request, 'importer_donnees.html', {'form': form})


# Fonction permettant à l'administrateur de gérer les localisations ajout des: régions, commune et ville
@login_required
@user_passes_test(lambda u: u.is_staff)
def gestion_localisation(request):
    region_form = RegionForm()
    commune_form = CommuneForm()
    ville_form = VilleForm()

    if request.method == 'POST':
        type_objet = request.POST.get('type_objet')
        if type_objet == 'region':
            region_form = RegionForm(request.POST)
            if region_form.is_valid():
                region_form.save()
                messages.success(request, "Région ajoutée.")
                return redirect('gestion_localisation')
            
        elif type_objet == 'commune':
            commune_form = CommuneForm(request.POST)
            if commune_form.is_valid():
                commune_form.save()
                messages.success(request, "Commune ajoutée.")
                return redirect('gestion_localisation')
            
        elif type_objet == 'ville':
            ville_form = VilleForm(request.POST)
            if ville_form.is_valid():
                ville_form.save()
                messages.success(request, "Ville ajoutée.")
                return redirect('gestion_localisation')
        
    return render(request, 'gestion_localisation.html', {
        'region_form': region_form,
        'ville_form': ville_form,
        'commune_form': commune_form,
        'regions': Region.objects.all().order_by('nom'),
        'communes': Commune.objects.select_related('region').order_by('nom'),
        'villes': Ville.objects.select_related('commune').order_by('nom'),
    })


# Fonction permettant à l'administrateur di'mporter en masse des régions/communes/ville
@login_required
@user_passes_test(lambda u: u.is_staff)
def importer_localisation_admin(request):
    if request.method == 'POST':
        form = ImportLocalisationForm(request.POST, request.FILES)
        if form.is_valid():
            fichier = request.FILES['fichier']
            try:
                df = pd.read_excel(fichier)
                df.columns = [str(c).strip() for c in df.columns]
                
                def get_val(row, possible_keys):
                    for k in possible_keys:
                        for col in row.index:
                            if str(col).strip().lower().replace(' ','_') == str(k).strip().lower().replace(' ', '_'):
                                val = row[col]
                                if pd.notna(val):
                                    return str(val).strip()
                    return None
                nb_regions, nb_communes, nb_villes = 0, 0, 0
                
                for index, row in df.iterrows():
                    nom_region = get_val(row, ['REGION'])
                    nom_commune = get_val(row, ['COMMUNE'])
                    nom_ville = get_val(row, ['VILLE'])
                    
                    if not nom_region or not nom_commune or not nom_ville:
                        continue
                    
                    region, created = Region.objects.get_or_create(nom=nom_region)
                    if created:
                        nb_regions += 1
                    
                    commune, created = Commune.objects.get_or_create(nom=nom_commune, region=region)
                    if created:
                        nb_communes += 1
                        
                    ville, created = Ville.objects.get_or_create(nom=nom_ville, commune=commune)
                    if created:
                        nb_villes += 1
                        
                messages.success(
                    request,
                    f"Importation réussie : {nb_regions} région(s), {nb_communes} commune(s), {nb_villes} ville(s) ajoutée(s)."
                )
                return redirect('gestion_localisation')
            
            except Exception as e:
                messages.error(request, f"Erreur lors de l'importation : {e}")
    else:
        form = ImportLocalisationForm()
    
    return render(request, 'importer_localisation.html', {'form':form})


# Fonction permettant à l'administrateur de supprimer une région
@login_required
@user_passes_test(lambda u: u.is_staff)
def supprimer_region(request, region_id):
    get_object_or_404(Region, id=region_id).delete()
    messages.success(request, 'Région supprimée.')
    return redirect('gestion_localisation')


# Fonction permettant à l'administrateur de supprimer une ville
@login_required
@user_passes_test(lambda u: u.is_staff)
def supprimer_ville(request, ville_id):
    get_object_or_404(Ville, id=ville_id).delete()
    messages.success(request, 'Ville supprimée.')
    return redirect('gestion_localisation')


# Fonction permettant à l'administrateur de supprimer une commune
@login_required
@user_passes_test(lambda u: u.is_staff)
def supprimer_Commune(request, Commune_id):
    get_object_or_404(Commune, id=Commune_id).delete()
    messages.success(request, 'Commune supprimée.')
    return redirect('gestion_localisation')


# Fonction permettant à l'administrateur de voir la liste des contrats
@login_required
@user_passes_test(lambda u: u.is_staff)
def liste_contrats_admin(request):
    quittances = Quittance.objects.select_related('contrat', 'contrat__user').order_by('contrat__user__last_name', 'contrat__user__first_name')
    query_police = request.GET.get('police', '').strip()
    query_nom = request.GET.get('nom', '').strip()
    query_type = request.GET.get('branche', '').strip()
    query_quittance = request.GET.get('quittance', '').strip()
    periode_option = request.GET.get('periode_option', '')
    date_effet_debut = request.GET.get('date_effet_debut', '')
    date_effet_fin = request.GET.get('date_effet_fin', '')

    today = timezone.localdate()
    if periode_option == 'auj':
        date_effet_debut = date_effet_fin = today.isoformat()
    elif periode_option == '7j':
        date_effet_debut = (today - timedelta(days=6)).isoformat()
        date_effet_fin = today.isoformat()
    elif periode_option == 'mois':
        date_effet_debut = today.replace(day=1).isoformat()
        date_effet_fin = today.isoformat()
    elif periode_option == 'trimestre':
        date_effet_debut = (today - timedelta(days=89)).isoformat()
        date_effet_fin = today.isoformat()
    elif periode_option == 'annee':
        date_effet_debut = today.replace(month=1, day=1).isoformat()
        date_effet_fin = today.isoformat()

    if query_police:
        quittances = quittances.filter(contrat__numero_police__icontains=query_police)
    if query_nom:
        quittances = quittances.filter(
            Q(contrat__user__last_name__icontains=query_nom) |
            Q(contrat__user__first_name__icontains=query_nom)
        )
    if query_type:
        quittances = quittances.filter(branche__icontains=query_type)
    if query_quittance:
        quittances = quittances.filter(numero_quittance__icontains=query_quittance)
    if date_effet_debut:
        quittances = quittances.filter(date_debut__gte=date.fromisoformat(date_effet_debut))
    if date_effet_fin:
        quittances = quittances.filter(date_debut__lte=date.fromisoformat(date_effet_fin))

    return render(request, 'liste_contrats.html', {
        'quittances': quittances,
        'query_police': query_police,
        'query_nom': query_nom,
        'query_type': query_type,
        'query_quittance': query_quittance,
        'periode_option': periode_option,
        'date_effet_debut': date_effet_debut,
        'date_effet_fin': date_effet_fin,
        'today': today,
    })
    
  
# Fonction permettant à l'administrateur de consulter le détail d'un dossier sinistre
@login_required
@user_passes_test(lambda u: u.is_staff)
def detail_sinistre_admin(request, sinistre_id):
    sinistre = get_object_or_404(Sinistre, id=sinistre_id)
    nb_non_lus = sinistre.messages.exclude(auteur=request.user).filter(lu=False).count()

    if request.method == 'POST' and 'contenu' in request.POST:
        Message.objects.create(sinistre=sinistre, auteur=request.user, contenu=request.POST.get('contenu'))
        return redirect('detail_sinistre_admin', sinistre_id=sinistre.id)

    nom_operateur = request.user.get_full_name() or request.user.username
    chef_est_traitant = sinistre.agent_traitant == nom_operateur

    sinistre.messages.exclude(auteur=request.user).filter(lu=False).update(lu=True)

    context = {
        'sinistre': sinistre,
        'chef_est_traitant': chef_est_traitant,
        'historique': sinistre.historique.all().order_by('date_changement'),
        'discussion': sinistre.messages.all().order_by('date_envoi'),
        'documents': sinistre.pieces.all(),
        'nb_non_lus': nb_non_lus,
    }
    return render(request, 'detail_sinistre_admin.html', context)


# Fonction permettant à l'administrateur de superviser les contrats par lot
@login_required
@user_passes_test(lambda u: u.is_staff)
def action_lot_contrats(request):
    if request.method != 'POST':
        return redirect('liste_contrats')

    ids = request.POST.getlist('selection')
    action = request.POST.get('action')

    if not ids:
        messages.warning(request, "Aucun contrat sélectionné.")
        return redirect('liste_contrats')

    if action == 'exporter':
        return exporter_contrats_admin(request)

    elif action == 'supprimer':
        supprimes = 0
        echecs = []

        with transaction.atomic():
            for assure_id in ids:
                try:
                    assure = Assure.objects.select_related('user').get(id=assure_id)
                    assure.user.delete()
                    supprimes += 1
                except Assure.DoesNotExist:
                    echecs.append(assure_id)

        if supprimes:
            messages.success(request, f"{supprimes} contrat(s) supprimé(s) avec succès.")
        if echecs:
            messages.error(
                request,
                f"{len(echecs)} contrat(s) introuvable(s) et n'ont pas pu être supprimés."
            )

    else:
        messages.error(request, "Action inconnue.")

    return redirect('liste_contrats')
    
    
# Fonction permettant à l'administrateur d'exporter des contrats avec des filtres
@login_required
@user_passes_test(lambda u: u.is_staff)
def exporter_contrats_admin(request):
    quittances = Quittance.objects.select_related('contrat', 'contrat__user').order_by('contrat__user__last_name')

    ids_selection = request.POST.getlist('selection')

    if ids_selection:
        quittances = quittances.filter(contrat__id__in=ids_selection)
    else:
        query_police = request.GET.get('police', '').strip()
        query_nom = request.GET.get('nom', '').strip()

        if query_police:
            quittances = quittances.filter(contrat__numero_police__icontains=query_police)
        if query_nom:
            quittances = quittances.filter(
                Q(contrat__user__last_name__icontains=query_nom) |
                Q(contrat__user__first_name__icontains=query_nom)
            )
            
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contrats"

    entetes =  ["N° Police", "Nom", "Prénom(s)", "Email", "Téléphone",
                "N° Quittance", "Branche", "Date effet", 
                "Date échéance", "Prime TTC(FCFA)", "Véhicule(s)"]

    ws.append(entetes)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for q in quittances:
        vehicules_liste = getattr(q.contrat, 'vehicules', None)
        vehicules = ", ".join(f"{v.immatriculation} ({v.marque})" for v in vehicules_liste.all()) if vehicules_liste else ''
        ws.append([
            q.contrat.numero_police,
            q.contrat.user.last_name,
            q.contrat.user.first_name,
            q.contrat.user.email,
            q.contrat.telephone or '',
            q.numero_quittance,
            q.branche or '',
            q.date_debut.strftime('%d/%m/%Y') if q.date_debut else '',
            q.date_fin.strftime('%d/%m/%Y') if q.date_fin else '',
            float(q.prime),
            vehicules,
        ])

    for col_cells in ws.columns:
        longueur = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(longueur + 2, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="contrats_fidelia.xlsx"'
    wb.save(response)
    return response


# Fonction permettant à l'administrateur de modifier un contrat
@login_required
@user_passes_test(lambda u: u.is_staff)
def modifier_contrat_admin(request, assure_id):
    assure = get_object_or_404(Assure, id=assure_id)
    if request.method == 'POST':
        form = AssureAdminForm(request.POST, instance=assure)
        if form.is_valid():
            form.save()
            messages.success(request, "Le contrat a été mis à jour avec succès.")
            return redirect('liste_contrats')
    else:
        form = AssureAdminForm(instance=assure)

    return render(request, 'modifier_contrat.html', {'form': form, 'assure': assure})


# Fonction permettant à l'administrateur de supprimer définitivement un sinistre
@login_required
@user_passes_test(lambda u: u.is_staff)
def supprimer_sinistre_admin(request, sinistre_id):
    sinistre = get_object_or_404(Sinistre, id=sinistre_id)

    if request.method == 'POST':
        numero = sinistre.numero_sinistre or f"#{sinistre.id}"
        sinistre.delete()
        messages.success(request, f"Le sinistre {numero} a été supprimé définitivement.")
        return redirect('supervision_sinistres')

    return render(request, 'confirmer_suppression_sinistre.html', {
        'sinistre': sinistre,
    })
    
    
# Fonction permettant à l'administrateur de modifier les informations d'un sinistre
@login_required
@user_passes_test(lambda u: u.is_staff)
def modifier_sinistre_admin(request, sinistre_id):
    sinistre = get_object_or_404(Sinistre, id=sinistre_id)

    if request.method == 'POST':
        form = ModifierSinistreAdminForm(request.POST, instance=sinistre)
        if form.is_valid():
            form.save()
            messages.success(request, f"Le sinistre {sinistre.numero_sinistre or sinistre.id} a été mis à jour.")
            return redirect('detail_sinistre_admin', sinistre_id=sinistre.id)
    else:
        form = ModifierSinistreAdminForm(instance=sinistre)

    return render(request, 'modifier_sinistre_admin.html', {'form': form, 'sinistre': sinistre})


# Fonction permettant à l'administrateur d'exporter les sinistres
@login_required
@user_passes_test(lambda u: u.is_staff)
def exporter_sinistres_admin(request):
    sinistres = Sinistre.objects.select_related('assure', 'vehicule', 'region', 'commune', 'ville').order_by('-date_declaration')

    # Cas 1 : export individuel (lien "Exporter" sur une ligne du tableau)
    id_unique = request.GET.get('id')
    # Cas 2 : export par lot (sélection cochée envoyée en POST)
    ids_selection = request.POST.getlist('selection')

    if id_unique:
        sinistres = sinistres.filter(id=id_unique)
    elif ids_selection:
        sinistres = sinistres.filter(id__in=ids_selection)
    else:
        # Cas 3 : export de la liste filtrée (mêmes filtres que supervision_sinistres)
        statut = request.GET.get('statut', '')
        nature = request.GET.get('nature', '')
        recherche = request.GET.get('q', '').strip()
        if statut:
            sinistres = sinistres.filter(statut=statut)
        if nature:
            sinistres = sinistres.filter(nature=nature)
        if recherche:
            sinistres = sinistres.filter(
                Q(numero_sinistre__icontains=recherche) | Q(assure__assure__numero_police__icontains=recherche)
            )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sinistres"

    entetes = ["N° Sinistre", "Nom", "Prénom(s)", "Email", "N° Police",
               "Immatriculation", "Nature", "Statut", "Agent traitant",
               "Montant estimé (FCFA)", "Prix retenu (FCFA)",
               "Date survenance", "Date déclaration",
               "Région", "Commune", "Ville", "Quartier", "Circonstances", "Dommages"]
    ws.append(entetes)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for s in sinistres:
        ws.append([
            s.numero_sinistre or '',
            s.assure.last_name,
            s.assure.first_name,
            s.assure.email,
            getattr(s.assure.assure, 'numero_police', '') if hasattr(s.assure, 'assure') else '',
            s.vehicule.immatriculation if s.vehicule else '',
            s.get_nature_display() if s.nature else '',
            s.get_statut_display(),
            s.agent_traitant or '',
            float(s.montant_estime or 0),
            float(s.prix_retenu) if s.prix_retenu is not None else '',
            s.date_survenance.strftime('%d/%m/%Y %H:%M') if s.date_survenance else '',
            s.date_declaration.strftime('%d/%m/%Y %H:%M') if s.date_declaration else '',
            s.region.nom if s.region else '',
            s.commune.nom if s.commune else '',
            s.ville.nom if s.ville else '',
            s.quartier or '',
            s.dommage,
            s.circonstances,
        ])

    for col_cells in ws.columns:
        longueur = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(longueur + 2, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sinistres_fidelia.xlsx"'
    wb.save(response)
    return response


# Fonction permettant à l'administrateur de faire des actions par lot sur les sinistres
@login_required
@user_passes_test(lambda u: u.is_staff)
def action_lot_sinistres(request):
    if request.method != 'POST':
        return redirect('supervision_sinistres')

    ids = request.POST.getlist('selection')
    action = request.POST.get('action')

    if not ids:
        messages.warning(request, "Aucun sinistre sélectionné.")
        return redirect('supervision_sinistres')

    if action == 'exporter':
        return exporter_sinistres_admin(request)

    elif action == 'supprimer':
        supprimes = 0
        echecs = []
        with transaction.atomic():
            for sinistre_id in ids:
                try:
                    sinistre = Sinistre.objects.get(id=sinistre_id)
                    sinistre.delete()
                    supprimes += 1
                except Sinistre.DoesNotExist:
                    echecs.append(sinistre_id)

        if supprimes:
            messages.success(request, f"{supprimes} sinistre(s) supprimé(s) avec succès.")
        if echecs:
            messages.error(request, f"{len(echecs)} sinistre(s) introuvable(s), non supprimés.")
    else:
        messages.error(request, "Action inconnue.")

    return redirect('supervision_sinistres')


# Fonction permettant à l'administrateur d'importer des sinistres
@login_required
@user_passes_test(lambda u: u.is_staff)
def importer_sinistres_admin(request):
    if request.method == 'POST':
        form = ImportSinistresForm(request.POST, request.FILES)
        if form.is_valid():
            fichier = request.FILES['fichier']
            try:
                df = pd.read_excel(fichier)
                df.columns = [str(c).strip() for c in df.columns]

                def _normalise_cle(valeur):
                    valeur = str(valeur).strip().lower()
                    valeur = unicodedata.normalize('NFKD', valeur)
                    valeur = ''.join(c for c in valeur if not unicodedata.combining(c))
                    valeur = valeur.replace('°', '').replace('n°', 'numero')
                    for caractere in ('_', '-', "'"):
                        valeur = valeur.replace(caractere, ' ')
                    return ' '.join(valeur.split())

                def get_val(row, possible_keys):
                    for k in possible_keys:
                        for col in row.index:
                            if _normalise_cle(col) == _normalise_cle(k):
                                val = row[col]
                                if pd.notna(val):
                                    return val
                    return None

                # Table de correspondance libellé -> code pour le statut (ex: "Soumis" -> "SOUMIS")
                statut_par_libelle = {
                    _normalise_cle(libelle): code for code, libelle in Sinistre.STATUS_CHOICES
                }
                statut_par_libelle.update({
                    _normalise_cle(code): code for code, _ in Sinistre.STATUS_CHOICES
                })

                def parser_statut(valeur):
                    if not valeur:
                        return 'SOUMIS'
                    return statut_par_libelle.get(_normalise_cle(valeur), 'SOUMIS')

                def parser_date(valeur):
                    if not valeur:
                        return None
                    if isinstance(valeur, datetime):
                        return valeur
                    horodatage = pd.to_datetime(str(valeur), dayfirst=True, errors='coerce')
                    if pd.isna(horodatage):
                        return None
                    return horodatage.to_pydatetime()

                crees, echecs = 0, []

                for index, row in df.iterrows():
                    try:
                        numero_police = get_val(row, ['NUMERO_POLICE', 'NUMERO POLICE', 'N° POLICE', 'N°POLICE', 'POLICE'])
                        if not numero_police:
                            echecs.append(f"Ligne {index + 2} : numéro de police manquant")
                            continue

                        assure = Assure.objects.filter(numero_police=numero_police).select_related('user').first()
                        if not assure:
                            echecs.append(f"Ligne {index + 2} : assuré introuvable ({numero_police})")
                            continue

                        immatriculation = get_val(row, ['IMMATRICULATION', 'IMMAT'])
                        vehicule = None
                        if immatriculation:
                            vehicule, _ = Vehicule.objects.get_or_create(
                                immatriculation=immatriculation,
                                defaults={
                                    'marque': get_val(row, ['MARQUE']) or '',
                                    'categorie': get_val(row, ['CATEGORIE']),
                                    'proprietaire': assure.user,
                                }
                            )
                        else:
                            echecs.append(f"Ligne {index + 2} : immatriculation manquante")
                            continue

                        def get_or_create_localisation(model, nom, **kwargs):
                            if not nom:
                                return None
                            obj, _ = model.objects.get_or_create(nom=str(nom).strip(), **kwargs)
                            return obj

                        region = get_or_create_localisation(Region, get_val(row, ['REGION']))
                        commune = get_or_create_localisation(Commune, get_val(row, ['COMMUNE']), region=region) if region else None
                        ville = get_or_create_localisation(Ville, get_val(row, ['VILLE']), commune=commune) if commune else None

                        nature_val = str(get_val(row, ['NATURE']) or '').strip().upper()[:1]
                        if nature_val not in ('C', 'M', 'X'):
                            nature_val = None

                        date_survenance = parser_date(get_val(row, ['DATE_SURVENANCE', 'DATE SURVENANCE']))
                        if not date_survenance:
                            echecs.append(f"Ligne {index + 2} : date de survenance manquante ou invalide")
                            continue

                        heure_val = get_val(row, ['HEURE_APPROXIMATIVE', 'HEURE'])
                        heure_approximative = parser_date(heure_val).time() if heure_val else date_survenance.time()

                        nom_conducteur = get_val(row, ['NOM_CONDUCTEUR', 'NOM CONDUCTEUR']) or (
                            f"{assure.user.first_name} {assure.user.last_name}".strip()
                        ) or assure.user.username

                        Sinistre.objects.create(
                            assure=assure.user,
                            vehicule=vehicule,
                            nom_conducteur=nom_conducteur,
                            contact_declarant=get_val(row, ['CONTACT_DECLARANT', 'CONTACT']),
                            immatriculation=immatriculation,
                            date_survenance=date_survenance,
                            heure_approximative=heure_approximative,
                            circonstances=get_val(row, ['CIRCONSTANCES']) or 'Non renseigné (import en masse).',
                            dommage=get_val(row, ['DOMMAGE']) or 'Non renseigné (import en masse).',
                            region=region,
                            commune=commune,
                            ville=ville,
                            quartier=get_val(row, ['QUARTIER']) or '',
                            nature=nature_val,
                            montant_estime=get_val(row, ['MONTANT_ESTIME', 'MONTANT']) or 0,
                            statut=parser_statut(get_val(row, ['STATUT'])),
                            agent_traitant=get_val(row, ['AGENT_TRAITANT', 'AGENT TRAITANT']) or '',
                        )
                        crees += 1

                    except Exception as e:
                        echecs.append(f"Ligne {index + 2} : {e}")

                if crees:
                    messages.success(request, f"{crees} sinistre(s) importé(s) avec succès.")
                if echecs:
                    messages.warning(request, f"{len(echecs)} ligne(s) ignorée(s) : " + " | ".join(echecs[:10]))

                return redirect('supervision_sinistres')

            except Exception as e:
                messages.error(request, f"Erreur lors de la lecture du fichier : {e}")
    else:
        form = ImportSinistresForm()

    return render(request, 'importer_sinistres.html', {'form': form})


# Fonction permettant à l'administrateur de supprimer un contrat
@login_required
@user_passes_test(lambda u: u.is_staff)
def supprimer_contrat_admin(request, assure_id):
    assure = get_object_or_404(Assure, id=assure_id)
    if request.method == 'POST':
        nom_complet = assure.user.get_full_name() or assure.user.username
        assure.user.delete()
        messages.success(request, f"Le contrat et le compte de {nom_complet} ont été supprimés.")
        return redirect('liste_contrats')

    return render(request, 'confirmer_suppression.html', {
        'objet_nom': f"Contrat {assure.numero_police} ({assure.user.get_full_name()})",
        'type_objet': 'contrat',
        'annuler_url': reverse('liste_contrats'),
        'confirmer_url': 'supprimer_contrat_admin',
        'objet_id': assure.id,
    })


# Fonction permettant à l'administrateur de gérer les différentes agences
@login_required
@user_passes_test(lambda u: u.is_staff)
def gestion_agences(request):
    if request.method == 'POST':
        form = AgenceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Point de vente ajouté avec succès.")
            return redirect('gestion_agences')
    else:
        form = AgenceForm()

    agences = Agence.objects.select_related('ville').order_by('nom')

    recherche = request.GET.get('q', '').strip()
    if recherche:
        agences = agences.filter(
            Q(nom__icontains=recherche) | Q(ville__nom__icontains=recherche)
        )

    return render(request, 'gestion_agences.html', {
        'form': form,
        'agences': agences,
        'recherche': recherche,
    })
        
    
#-------------------------------------------------------------------
#                   INDEMNISATION
#-------------------------------------------------------------------


# Gère l'enregistrement de l'indemnisation d'un sinistre.
@login_required
def indemniser_sinistre(request, sinistre_id):
    sinistre = get_object_or_404(Sinistre, id=sinistre_id)
    
    if request.method == 'POST':
        # Passer le sinistre au formulaire pour la validation du montant
        form = IndemnisationForm(request.POST, sinistre=sinistre)
        
        if form.is_valid():
            data = form.cleaned_data
            
            # Mise à jour des champs sur l'instance de Sinistre
            sinistre.beneficiaire_nom = data['beneficiaire_nom']
            sinistre.beneficiaire_prenoms = data['beneficiaire_prenoms']
            sinistre.beneficiaire_telephone = data['beneficiaire_telephone']
            sinistre.numero_cheque = data['numero_cheque']
            sinistre.banque_cheque = data['banque_cheque']
            sinistre.montant_cheque = data['montant_cheque']
            sinistre.date_emission_cheque = data['date_emission_cheque']
            
            # Changement de statut du dossier
            sinistre.statut = 'INDEMNISE'  # Ajustez selon vos choix de statut
            sinistre.save()
            
            messages.success(
                request, 
                f"L'indemnisation par chèque N° {data['numero_cheque']} pour le dossier {sinistre.id} a été enregistrée avec succès."
            )
            return redirect('detail_sinistre_agent', sinistre_id=sinistre.id)
    else:
        # Pré-remplissage des données initiales du bénéficiaire (Assuré)
        initial_data = {}
        if hasattr(sinistre, 'vehicule') and sinistre.vehicule.proprietaire:
            user_assure = sinistre.vehicule.proprietaire
            initial_data['beneficiaire_nom'] = user_assure.last_name
            initial_data['beneficiaire_prenoms'] = user_assure.first_name
            
            # Si vous avez un profil Assure lié à User
            if hasattr(user_assure, 'assure'):
                initial_data['beneficiaire_telephone'] = user_assure.assure.telephone

        form = IndemnisationForm(initial=initial_data, sinistre=sinistre)

    return render(request, 'sinistres/indemniser.html', {
        'form': form,
        'sinistre': sinistre
    })


# Fonction permettant l'émission d'un chèque
@login_required
def emettre_cheque(request, sinistre_id):
    operateur = get_profil_operateur(request.user)
    if not operateur:
        return redirect('accueil_assure')
    
    sinistre = get_object_or_404(Sinistre, pk=sinistre_id)
    detail_url_name = get_url_detail_dossier(request.user)
    
    if request.method == 'POST':
        numero_cheque = request.POST.get('numero_cheque', '').strip()
        banque_cheque = request.POST.get('banque_cheque')
        montant = request.POST.get('montant')

        # Le numéro vient du chèque papier déjà remis à l'agent : on vérifie juste l'unicité
        if Paiement.objects.filter(numero_cheque=numero_cheque).exists():
            messages.error(
                request,
                f"Le numéro de chèque « {numero_cheque} » est déjà enregistré dans le système. "
                "Vérifiez le numéro inscrit sur le chèque."
            )
            return redirect('emettre_cheque', sinistre_id=sinistre.id)
        beneficiaire_nom = request.POST.get('beneficiaire_nom')
        beneficiaire_prenoms = request.POST.get('beneficiaire_prenoms')
        beneficiaire_telephone = request.POST.get('beneficiaire_telephone')
        date_emission = request.POST.get('date_emission')

        # Vérification : la date d'émission ne peut pas être dans le futur
        try:
            date_emission_obj = datetime.strptime(date_emission, '%Y-%m-%d').date()
        except (TypeError, ValueError):
            messages.error(request, "Date d'émission invalide.")
            return redirect('emettre_cheque', sinistre_id=sinistre.id)
        
        if date_emission_obj > timezone.now().date():
            messages.error(request, "La date d'émission du chèque ne peut pas être dans le futur.")
            return redirect('emettre_cheque', sinistre_id=sinistre.id)
        
        
        # Vérification : le montant ne doit pas dépasser le reste à payer
        try:
            montant_decimal = Decimal(montant)
        except (TypeError, ValueError, InvalidOperation):
            messages.error(request, "Montant invalide.")
            return redirect('emettre_cheque', sinistre_id=sinistre.id)

        if montant_decimal > sinistre.reste_a_payer:
            messages.error(
                request,
                f"Le montant du chèque ({montant_decimal} FCFA) dépasse le reste à payer "
                f"({sinistre.reste_a_payer} FCFA)."
            )
            return redirect('emettre_cheque', sinistre_id=sinistre.id)

        try:
            paiement = Paiement.objects.create(
                sinistre=sinistre,
                numero_cheque=numero_cheque,
                banque_cheque=banque_cheque,
                montant=montant_decimal,
                beneficiaire_nom=beneficiaire_nom,
                beneficiaire_prenoms=beneficiaire_prenoms,
                beneficiaire_telephone=beneficiaire_telephone,
                date_emission=date_emission,
                statut='EMIS'
            )
        except IntegrityError:
            messages.error(
                request,
                f"Le numéro de chèque « {numero_cheque} » est déjà enregistré (doublon)."
            )
            return redirect('emettre_cheque', sinistre_id=sinistre.id)
        
       
        HistoriqueSinistre.objects.create(
            sinistre=sinistre,
            statut='Chèque émis',
            commentaires=f"Émission d'un chèque de {montant_decimal} FCFA (N° {numero_cheque}) tiré sur {banque_cheque} pour {beneficiaire_prenoms} {beneficiaire_nom}.",
            auteur=request.user
        )
        
        return redirect('detail_sinistre_agent', sinistre_id=sinistre.pk)

    return render(request, 'emettre_cheque.html', {'sinistre': sinistre, 'detail_url_name': detail_url_name})


# Fonction permettant la modification du statut d'un chèque
@login_required
def modifier_statut_cheque(request, paiement_id):
    paiement = get_object_or_404(Paiement, pk=paiement_id)
    nouveau_statut = request.POST.get('statut')
    
    if nouveau_statut == 'RETIRE':
        return redirect('marquer_cheque_retire', paiement_id=paiement_id)

    if nouveau_statut in dict(Paiement.STATUT_PAIEMENT).keys():
        paiement.statut = nouveau_statut
        paiement.save()

        HistoriqueSinistre.objects.create(
            sinistre=paiement.sinistre,
            statut="Mise à jour chèque",
            commentaires=f"Le statut du chèque N° {paiement.numero_cheque} est passé à: {paiement.get_statut_display()}.",
            auteur=request.user
        )
        
        if nouveau_statut == 'DISPONIBLE':
            Message.objects.create(
                sinistre=paiement.sinistre,
                auteur=request.user,
                contenu=f"Votre chèque N° {paiement.numero_cheque} d'un montant de {paiement.montant} FCFA est prêt pour retrait. Merci de vous présenter avec votre pièce d'identité."
            )
            
        messages.success(request, f"Le statut du chèque {paiement.numero_cheque} a été mis à jour avec succès.")
    else:
        messages.error(request, "Statut invalide.")

    return redirect('detail_sinistre_agent', sinistre_id=paiement.sinistre.pk)


# Fonction permettant de définir l'état définitif d'un chèque
@login_required
def marquer_cheque_retire(request, paiement_id):
    paiement = get_object_or_404(Paiement, pk=paiement_id)

    if request.method == 'POST':
        form = RetraitChequeForm(request.POST, request.FILES)
        if form.is_valid():
            paiement.statut = 'RETIRE'
            paiement.nom_retirant = form.cleaned_data['nom_retirant']
            paiement.type_piece_retirant = form.cleaned_data['type_piece_retirant']
            paiement.numero_piece_retirant = form.cleaned_data['numero_piece_retirant']
            paiement.piece_identite_retirant = form.cleaned_data['piece_identite_retirant']
            paiement.save()

            if paiement.piece_identite_retirant:
                PieceJointe.objects.create(
                    sinistre=paiement.sinistre,
                    fichier=paiement.piece_identite_retirant,
                )

            HistoriqueSinistre.objects.create(
                sinistre = paiement.sinistre,
                statut='Chèque retiré',
                commentaires=(
                    f"Chèque N° {paiement.numero_cheque} retiré par "
                    f"{paiement.nom_retirant or (paiement.beneficiaire_prenoms + ' ' + paiement.beneficiaire_nom)}"
                    f"(pièce: {paiement.get_type_piece_retirant_display()} N° {paiement.numero_piece_retirant})."
                ),
                auteur=request.user,
            )
            messages.success(request, f"Chèque {paiement.numero_cheque} marqué comme retiré.")
            return redirect('detail_sinistre_agent', sinistre_id=paiement.sinistre.pk)
    else:
         form = RetraitChequeForm()

    return render(request, 'marquer_cheque_retire.html', {'form': form, 'paiement': paiement})


# Première étape de récupération du mot de passe (vérification identifiant/téléphone).
def mot_de_passe_oublie_etape1(request):
    if request.method == 'POST':
        form = MotDePasseOublieForm(request.POST)
        if form.is_valid():
            identifiant = form.cleaned_data['identifiant'].strip()
            telephone = form.cleaned_data['telephone'].strip()
            profil = get_profil_par_identifiant(identifiant)
            telephone_profil = getattr(profil, 'telephone', None) if profil else None

            if profil and telephone_profil and telephone_profil.strip() == telephone:
                request.session['reset_user_id'] = profil.user.id
                return redirect('mot_de_passe_oublie_etape2')
            messages.error(request, "Identifiant ou numéro de téléphone incorrect.")
    else:
        form = MotDePasseOublieForm()
    return render(request, 'mot_de_passe_oublie_etape1.html', {'form': form})


# Deuxième étape : Saisie du nouveau mot de passe après vérification.
def mot_de_passe_oublie_etape2(request):
    user_id = request.session.get('reset_user_id')
    if not user_id:
        messages.error(request, "Veuillez d'abord renseigner votre identifiant et votre numéro de téléphone.")
        return redirect('mot_de_passe_oublie_etape1')
    
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        form = SetPasswordForm(user, request.POST)
        if form.is_valid():
            form.save()
            del request.session['reset_user_id']

            profil = (
                getattr(user, 'assure', None)
                or getattr(user, 'agent', None)
                or getattr(user, 'chef', None)
            )

            if profil:
                if hasattr(profil, 'doit_changer_mot_de_passe'):
                    profil.doit_changer_mot_de_passe = False
                    profil.save(update_fields=['doit_changer_mot_de_passe'])

                # Réactivation du compte : on remet le compteur de tentatives
                # et le blocage à zéro suite à la réinitialisation du mot de passe.
                reinitialiser_tentatives(profil)

            messages.success(request, "Votre mot de passe a été réinitialisé.")
            return redirect('login')
    else:
        form = SetPasswordForm(user)

    return render(request, 'mot_de_passe_oublie_etape2.html', {'form': form})


#-------------------------------------------------------------------
#                   GENERAL
#-------------------------------------------------------------------


# Fonction permettant à tous les utilisateurs de modifier leur mot de passe
@login_required
def changer_mot_de_passe_en_cours_admin(request):
    if request.method == 'POST':
        form = StylePasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, form.user)
            messages.success(request, "Mot de passe modifié avec succès.")
            return redirect('redirection_login')
    else:
        form = StylePasswordChangeForm(request.user)
    return render(request, 'changer_mot_de_passe_en_cours_admin.html', {'form': form})



@login_required
def changer_mot_de_passe_en_cours_agent(request):
    if request.method == 'POST':
        form = StylePasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, form.user)
            messages.success(request, "Mot de passe modifié avec succès.")
            return redirect('redirection_login')
    else:
        form = StylePasswordChangeForm(request.user)
    return render(request, 'changer_mot_de_passe_en_cours_agent.html', {'form': form})


@login_required
def changer_mot_de_passe_en_cours_chef(request):
    if request.method == 'POST':
        form = StylePasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, form.user)
            messages.success(request, "Mot de passe modifié avec succès.")
            return redirect('redirection_login')
    else:
        form = StylePasswordChangeForm(request.user)
    return render(request, 'changer_mot_de_passe_en_cours_chef.html', {'form': form})


@login_required
def changer_mot_de_passe_en_cours_assure(request):
    if request.method == 'POST':
        form = StylePasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, form.user)
            messages.success(request, "Mot de passe modifié avec succès.")
            return redirect('redirection_login')
    else:
        form = StylePasswordChangeForm(request.user)
    return render(request, 'changer_mot_de_passe_en_cours_assure.html', {'form': form})